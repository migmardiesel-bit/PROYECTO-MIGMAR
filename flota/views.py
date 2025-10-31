# flota/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.views.decorators.http import require_POST
from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.utils.dateparse import parse_date
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.core.exceptions import PermissionDenied
from decimal import Decimal
from datetime import date, timedelta, datetime
from django.views.generic.edit import FormView
from django.db import transaction
from django.contrib import messages
from django.views.decorators.http import require_POST # <-- AÑADIR ESTE IMPORT
from django.forms import formset_factory, inlineformset_factory, modelformset_factory
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo # <-- AÑADIR ESTA LÍNEA
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule
from django.http import JsonResponse
from .models import *
from .forms import *
from django.utils import timezone
from .models import Unidad, ChecklistCorreccion
import openpyxl # <--- AÑADE ESTE IMPORT
from openpyxl.styles import Font, Alignment # <--- AÑADE ESTE IMPORT
# --- NUEVOS IMPORTS PARA PDF ---
from django.http import HttpResponse
import json
from django.template.loader import render_to_string
try:
    from weasyprint import HTML
except ImportError:
    HTML = None # Permite que el servidor se ejecute incluso si WeasyPrint no está instalado

from django.http import JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin # Asegúrate que UserPassesTestMixin esté importado
from .alerts import send_on_demand_status_report # <-- NUEVO IMPORT
from .models import AsignacionRevision # <-- Y todos los demás modelos que uses
# === IMPORTS ADICIONALES PARA CARGA MANUAL A S3 ===
import boto3
from botocore.exceptions import BotoCoreError, NoCredentialsError
from django.conf import settings
import os
from .utils import recalcular_costos_cargas_diesel, recalcular_costos_cargas_urea
# ===================================================


# ==============================================================================
# === NUEVAS FUNCIONES AUXILIARES PARA GESTIONAR ARCHIVOS EN S3 MANUALMENTE ===
# ==============================================================================

def _subir_archivo_a_s3(archivo_obj, s3_ruta_relativa):
    """
    Sube un archivo a S3.
    - `s3_ruta_relativa` es la ruta SIN 'media/' (ej: 'flota/checklists/foto.jpg').
    - Devuelve la misma ruta relativa si tiene éxito, para guardarla en la DB.
    """
    try:
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
        
        # Boto3 necesita la ruta completa (Key) dentro del bucket
        full_s3_path = f"{settings.AWS_MEDIA_LOCATION}/{s3_ruta_relativa}"

        # Asegurarse de que el archivo esté al inicio
        archivo_obj.seek(0)
        
        s3_client.upload_fileobj(
            archivo_obj,
            settings.AWS_STORAGE_BUCKET_NAME,
            full_s3_path
        )
        return s3_ruta_relativa
        
    except (BotoCoreError, NoCredentialsError, Exception) as e:
        print(f"Error al subir el archivo a S3: {e}")
        return None

def _eliminar_archivo_de_s3(ruta_completa_s3):
    """
    Elimina un archivo de S3.
    - `ruta_completa_s3` es la ruta que Django provee (ej: 'media/entradas/foto.jpg'),
      que es lo que Boto3 necesita como 'Key'.
    """
    if not ruta_completa_s3:
        return
    try:
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
        s3_client.delete_object(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=str(ruta_completa_s3)
        )
    except (BotoCoreError, NoCredentialsError, Exception) as e:
        print(f"Error al eliminar archivo antiguo de S3: {e}")


# ===================================================================
# 1. PERMISSIONS AND ROLE LOGIC
# ===================================================================

def es_admin(user):
    return user.is_authenticated and user.groups.filter(name='Administrador').exists()

def es_tecnico(user):
    return user.is_authenticated and user.groups.filter(name='Tecnico').exists()

def es_encargado(user):
    return user.is_authenticated and user.groups.filter(name='Encargado').exists()
class AdminOrEncargadoRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """
    Permite el acceso a usuarios que son Administradores O Encargados.
    """
    def test_func(self):
        return es_admin(self.request.user) or es_encargado(self.request.user)
def puede_iniciar_proceso(user):
    return es_tecnico(user) or es_encargado(user)

@login_required
def home_dispatcher_view(request):
    """Redirects each user to their home page based on their role."""
    if es_admin(request.user):
        return redirect('dashboard')
    elif es_encargado(request.user):
        return redirect('encargado-pendientes-list')
    elif es_tecnico(request.user):
        return redirect('tecnico-seleccionar-unidad')
    else:
        raise PermissionDenied("You do not have a role assigned to access this application.")

class AdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return es_admin(self.request.user)

class IniciaProcesoRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return puede_iniciar_proceso(self.request.user)

class EncargadoRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return es_encargado(self.request.user)

# ===================================================================
# 2. ADMINISTRATOR VIEWS (DASHBOARD, CRUDs, APIs)
# ===================================================================

# --- Dashboards and APIs ---
class AdminDashboardView(AdminRequiredMixin, ListView):
    model = Unidad
    template_name = 'dashboard.html'
    context_object_name = 'unidades'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # --- CÁLCULO DE INVENTARIO ---
        
        # 1. Compras totales
        total_diesel_comprado = CompraSuministro.objects.filter(tipo_suministro='DIESEL').aggregate(total=Sum('cantidad'))['total'] or 0
        total_urea_comprada = CompraSuministro.objects.filter(tipo_suministro='UREA').aggregate(total=Sum('cantidad'))['total'] or 0
        total_aceite_comprado = CompraSuministro.objects.filter(tipo_suministro='ACEITE').aggregate(total=Sum('cantidad'))['total'] or 0

        # 2. Consumos totales
        consumo_diesel_total = CargaDiesel.objects.aggregate(total=Sum('lts_diesel'))['total'] or 0
        consumo_thermo_total = CargaDiesel.objects.aggregate(total=Sum('lts_thermo'))['total'] or 0
        consumo_urea_total = CargaUrea.objects.aggregate(total=Sum('litros_cargados'))['total'] or 0
        consumo_aceite_total = CargaAceite.objects.aggregate(total=Sum('cantidad'))['total'] or 0

        # 3. Suma de todos los ajustes (entradas y salidas)
        ajustes_diesel = AjusteInventario.objects.filter(tipo_suministro='DIESEL').aggregate(total=Sum('cantidad'))['total'] or 0
        ajustes_urea = AjusteInventario.objects.filter(tipo_suministro='UREA').aggregate(total=Sum('cantidad'))['total'] or 0
        ajustes_aceite = AjusteInventario.objects.filter(tipo_suministro='ACEITE').aggregate(total=Sum('cantidad'))['total'] or 0

        # 4. Cálculo final del inventario: (Compras - Consumos + Ajustes)
        context['diesel_restante'] = total_diesel_comprado - (consumo_diesel_total + consumo_thermo_total) + ajustes_diesel
        context['urea_restante'] = total_urea_comprada - consumo_urea_total + ajustes_urea
        context['aceite_restante'] = total_aceite_comprado - consumo_aceite_total + ajustes_aceite
        context['alertas_activas'] = AlertaInventario.objects.filter(activa=True)
        # --- REPORTE DE CONSUMO POR UNIDAD ---
        
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        if start_date_str and end_date_str:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
        else:
            today = date.today()
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            
        context['start_date'] = start_date
        context['end_date'] = end_date
        
        unidades = Unidad.objects.all()
        consumo_diesel = CargaDiesel.objects.filter(fecha__date__range=[start_date, end_date]).values('unidad_id').annotate(total_motor=Sum('lts_diesel'), total_thermo=Sum('lts_thermo'), total_costo=Sum('costo'))
        consumo_urea = CargaUrea.objects.filter(fecha__date__range=[start_date, end_date]).values('unidad_id').annotate(total_urea=Sum('litros_cargados'))
        consumo_aceite = CargaAceite.objects.filter(fecha__date__range=[start_date, end_date]).values('unidad_id').annotate(total_aceite=Sum('cantidad'))
        
        diesel_map = {item['unidad_id']: item for item in consumo_diesel}
        urea_map = {item['unidad_id']: item for item in consumo_urea}
        aceite_map = {item['unidad_id']: item for item in consumo_aceite}
        
        consumo_por_unidad = []
        for unidad in unidades:
            diesel_data = diesel_map.get(unidad.id, {})
            urea_data = urea_map.get(unidad.id, {})
            aceite_data = aceite_map.get(unidad.id, {})
            consumo_por_unidad.append({
                'unidad': unidad, 'motor': diesel_data.get('total_motor', 0) or 0, 'thermo': diesel_data.get('total_thermo', 0) or 0, 
                'urea': urea_data.get('total_urea', 0) or 0, 'aceite': aceite_data.get('total_aceite', 0) or 0, 'costo': diesel_data.get('total_costo', 0) or 0,
            })
        
        consumo_por_unidad.sort(key=lambda item: item['motor'] + item['thermo'], reverse=True)
        context['consumo_por_unidad'] = consumo_por_unidad
        
        return context
    
class DashboardGraficasView(AdminRequiredMixin, ListView):
    model = Unidad
    template_name = 'dashboard_graficas.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Graphics logic from your old file
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        if start_date_str and end_date_str:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
        else:
            today = date.today()
            last_sunday = today - timedelta(days=(today.weekday() + 1))
            start_date = last_sunday - timedelta(days=6)
            end_date = last_sunday
            
        context['start_date'] = start_date
        context['end_date'] = end_date

        total_diesel_comprado_rango = CompraSuministro.objects.filter(
            tipo_suministro='DIESEL',
            fecha_compra__range=[start_date, end_date]
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        context['total_diesel_comprado_rango'] = total_diesel_comprado_rango

        consumo_diesel = CargaDiesel.objects.filter(fecha__date__range=[start_date, end_date]) \
            .values('unidad_id') \
            .annotate(total_motor=Sum('lts_diesel'), total_thermo=Sum('lts_thermo'))
        
        diesel_map = {item['unidad_id']: item for item in consumo_diesel}

        unidades = Unidad.objects.all()
        consumo_por_unidad = []
        for unidad in unidades:
            diesel_data = diesel_map.get(unidad.id, {})
            motor = diesel_data.get('total_motor', 0) or 0
            thermo = diesel_data.get('total_thermo', 0) or 0
            
            if motor > 0 or thermo > 0:
                consumo_por_unidad.append({
                    'unidad': unidad.nombre, 
                    'motor': motor, 
                    'thermo': thermo
                })
        
        motor_chart_data = sorted([item for item in consumo_por_unidad if item['motor'] > 0], key=lambda x: x['motor'], reverse=True)
        context['motor_chart_data'] = motor_chart_data

        thermo_chart_data = sorted([item for item in consumo_por_unidad if item['thermo'] > 0], key=lambda x: x['thermo'], reverse=True)
        context['thermo_chart_data'] = thermo_chart_data

        total_motor_rango = sum(item['motor'] for item in consumo_por_unidad)
        total_thermo_rango = sum(item['thermo'] for item in consumo_por_unidad)
        context['total_motor_rango'] = total_motor_rango
        context['total_thermo_rango'] = total_thermo_rango
        context['total_general_rango'] = total_motor_rango + total_thermo_rango

        return context

@login_required
def get_unidad_tipo_api(request):
    unidad_id = request.GET.get('unidad_id')
    if not unidad_id: return JsonResponse({'error': 'Unit ID not provided'}, status=400)
    try:
        unidad = Unidad.objects.get(pk=unidad_id)
        return JsonResponse({'tipo': unidad.tipo})
    except Unidad.DoesNotExist:
        return JsonResponse({'error': 'Unit not found'}, status=404)

@login_required
def search_unidades_api(request):
    term = request.GET.get('term', '')
    unidades = Unidad.objects.filter(nombre__icontains=term)[:15]
    results = [{'id': u.id, 'text': u.nombre} for u in unidades]
    return JsonResponse({'results': results})

@login_required
def search_tecnicos_api(request):
    term = request.GET.get('term', '')
    tecnicos = User.objects.filter(Q(groups__name='Tecnico') & (Q(username__icontains=term) | Q(first_name__icontains=term) | Q(last_name__icontains=term))).distinct()[:15]
    results = [{'id': t.id, 'text': t.get_full_name() or t.username} for t in tecnicos]
    return JsonResponse({'results': results})

# --- Unit CRUD ---
class UnidadListView(AdminRequiredMixin, ListView):
    model = Unidad
    template_name = 'generic_list.html'
    paginate_by = 10
    def get_queryset(self):
        queryset = super().get_queryset().order_by('nombre')
        search_query = self.request.GET.get('q')
        tipo_filter = self.request.GET.get('tipo_unidad')
        if search_query:
            queryset = queryset.filter(nombre__icontains=search_query)
        if tipo_filter:
            queryset = queryset.filter(tipo=tipo_filter)
        return queryset
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({'titulo': 'Unidades', 'url_crear': 'unidad-create', 'headers': ['#', 'Nombre', 'Marca/Modelo', 'Placas', 'KM Actual', 'Tipo'], 'url_detail_name': 'unidad-detail', 'url_update_name': 'unidad-update', 'url_delete_name': 'unidad-delete'})
        context['search_query'] = self.request.GET.get('q', '')
        context['selected_tipo'] = self.request.GET.get('tipo_unidad', '')
        
        # ========= INICIO DEL CÓDIGO A AGREGAR =========
        # Habilita el botón de exportación general en la plantilla
        context['show_general_export_button'] = True
        # Define la URL a la que apuntará el botón
        context['export_url_name'] = 'unidades-export-excel'
        # ========= FIN DEL CÓDIGO A AGREGAR =========
        
        return context

class UnidadCreateView(AdminRequiredMixin, CreateView):
    model = Unidad
    form_class = UnidadForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('unidad-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Add New Unit'
        return context

class UnidadUpdateView(AdminRequiredMixin, UpdateView):
    model = Unidad
    form_class = UnidadForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('unidad-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f"Editar Unidad: {self.object.nombre}"
        return context

class UnidadDeleteView(AdminRequiredMixin, DeleteView):
    model = Unidad
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('unidad-list')
    def form_valid(self, form):
        messages.success(self.request, f"Unidad '{self.object.nombre}' successfully deleted.")
        return super().form_valid(form)

class UnidadDetailRendimientoView(AdminRequiredMixin, DetailView):
    model = Unidad
    template_name = 'unidad_rendimiento_detail.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unidad = self.get_object()
        context['cargas_diesel'] = CargaDiesel.objects.filter(unidad=unidad).order_by('-fecha')
        context['cargas_urea'] = CargaUrea.objects.filter(unidad=unidad).order_by('-fecha')
        return context

# --- Operator CRUD ---
class OperadorListView(AdminRequiredMixin, ListView):
    model = Operador
    template_name = 'generic_list.html'
    paginate_by = 10
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({'titulo': 'Operadores', 'url_crear': 'operador-create', 'headers': ['#', 'Nombre Completo', 'Teléfono'], 'url_detail_name': 'operador-detail', 'url_update_name': 'operador-update', 'url_delete_name': 'operador-delete'})
        return context

class OperadorCreateView(AdminRequiredMixin, CreateView):
    model = Operador
    form_class = OperadorForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('operador-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Register New Operator'
        return context

class OperadorUpdateView(AdminRequiredMixin, UpdateView):
    model = Operador
    form_class = OperadorForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('operador-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f"Edit Operator: {self.object}"
        return context

class OperadorDetailView(AdminRequiredMixin, DetailView):
    model = Operador
    template_name = 'operador_detail.html'

class OperadorDeleteView(AdminRequiredMixin, DeleteView):
    model = Operador
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('operador-list')
    def form_valid(self, form):
        messages.success(self.request, f"Operator '{self.object}' successfully deleted.")
        return super().form_valid(form)

# --- Diesel Load CRUD ---
class CargaDieselListView(AdminRequiredMixin, ListView):
    model = CargaDiesel
    template_name = 'generic_list.html'
    paginate_by = 25

    def get_queryset(self):
        """
        Queryset modificado para aceptar filtros y optimizar
        la carga de fotos de odómetro.
        """
        # --- MODIFICACIÓN: Añadir select_related ---
        queryset = super().get_queryset().select_related(
            'unidad', 
            'procesocarga__checklist' # Optimiza la búsqueda de la foto del odómetro
        ).order_by('-fecha')
        # --- FIN MODIFICACIÓN ---
        
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        unidad_id = self.request.GET.get('unidad')

        if start_date_str and end_date_str:
            # Filtra los registros que están dentro del rango de fechas.
            queryset = queryset.filter(fecha__date__range=[date.fromisoformat(start_date_str), date.fromisoformat(end_date_str)])
        
        if unidad_id:
            # Filtra por el ID de la unidad seleccionada.
            queryset = queryset.filter(unidad_id=unidad_id)
            
        return queryset

    def get_context_data(self, **kwargs):
        """
        Contexto modificado para añadir los nuevos encabezados de la tabla
        y para devolver los valores de los filtros a la plantilla.
        """
        context = super().get_context_data(**kwargs)
        
        # Nuevos encabezados para la tabla
        context.update({
            'titulo': 'Cargas de Diésel', 
            'url_crear': 'cargadiesel-create', 
            'headers': ['#', 'Unidad', 'Fecha', 'Lts Motor', 'Lts Thermo', 'Total Consumido', 'Costo'], # <-- ENCABEZADOS ACTUALIZADOS
            'url_update_name': 'cargadiesel-update', 
            'url_delete_name': 'cargadiesel-delete'
        })

        # Devuelve los valores de los filtros para que se mantengan en los campos del formulario
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        
        # Si se seleccionó una unidad, la obtenemos para mostrarla en el campo Select2
        if selected_unidad_id := self.request.GET.get('unidad'):
            context['selected_unidad'] = Unidad.objects.filter(pk=selected_unidad_id).first()
        
        # ========= INICIO DEL CÓDIGO A AGREGAR =========
        # Habilita el botón de exportación general en la plantilla
        context['show_general_export_button'] = True
        # Define la URL a la que apuntará el botón
        context['export_url_name'] = 'cargadiesel-export-excel'
        # ========= FIN DEL CÓDIGO A AGREGAR =========
            
        return context

class CargaDieselCreateView(AdminRequiredMixin, CreateView):
    model = CargaDiesel
    form_class = CargaDieselForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargadiesel-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Carga de diesel'
        return context
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

class CargaDieselUpdateView(AdminRequiredMixin, UpdateView):
    model = CargaDiesel
    form_class = CargaDieselForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargadiesel-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f"Edicion Diesel {self.object.fecha.strftime('%d/%m/%Y')}"
        return context
    
class CargaDieselDeleteView(AdminRequiredMixin, DeleteView):
    model = CargaDiesel
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('cargadiesel-list')

class CargaAceiteListView(AdminRequiredMixin, ListView):
    model = CargaAceite
    template_name = 'generic_list.html'
    paginate_by = 25

    # --- MÉTODO MODIFICADO ---
    def get_queryset(self):
        queryset = super().get_queryset().order_by('-fecha')
        
        # Lógica de filtrado añadida
        unidad_id = self.request.GET.get('unidad')
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        if unidad_id:
            queryset = queryset.filter(unidad_id=unidad_id)
        if start_date_str and end_date_str:
            queryset = queryset.filter(fecha__date__range=[date.fromisoformat(start_date_str), date.fromisoformat(end_date_str)])
            
        return queryset

    # --- MÉTODO MODIFICADO ---
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Cargas de Aceite', 
            'url_crear': 'cargaaceite-create', 
            'headers': ['#', 'Unidad', 'Fecha', 'Cantidad', 'Motivo'], 
            'url_update_name': 'cargaaceite-update', 
            'url_delete_name': 'cargaaceite-delete'
        })
        # Pasamos los valores de los filtros a la plantilla
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        if selected_unidad_id := self.request.GET.get('unidad'):
            context['selected_unidad'] = Unidad.objects.filter(pk=selected_unidad_id).first()
        return context

class CargaAceiteCreateView(AdminRequiredMixin, CreateView):
    model = CargaAceite
    form_class = CargaAceiteForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargaaceite-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Registrar Carga de Aceite'
        return context

class CargaAceiteUpdateView(AdminRequiredMixin, UpdateView):
    model = CargaAceite
    form_class = CargaAceiteForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargaaceite-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f"Editar Aceite {self.object.fecha.strftime('%d/%m/%Y')}"
        return context

class CargaAceiteDeleteView(AdminRequiredMixin, DeleteView):
    model = CargaAceite
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('cargaaceite-list')

# --- Urea Load CRUD ---
class CargaUreaListView(AdminRequiredMixin, ListView):
    model = CargaUrea
    template_name = 'generic_list.html'
    paginate_by = 25

    def get_queryset(self):
        queryset = super().get_queryset().order_by('-fecha')
        
        unidad_id = self.request.GET.get('unidad')
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        if unidad_id:
            queryset = queryset.filter(unidad_id=unidad_id)
        if start_date_str and end_date_str:
            queryset = queryset.filter(fecha__date__range=[date.fromisoformat(start_date_str), date.fromisoformat(end_date_str)])

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Cargas de Urea', 
            'url_crear': 'cargaurea-create', 
            # --- LÍNEA CORREGIDA ---
            'headers': ['#', 'Unidad', 'Fecha', 'Litros', 'Costo'], 
            'url_update_name': 'cargaurea-update', 
            'url_delete_name': 'cargaurea-delete'
        })
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        if selected_unidad_id := self.request.GET.get('unidad'):
            context['selected_unidad'] = Unidad.objects.filter(pk=selected_unidad_id).first()
        return context

class CargaUreaCreateView(AdminRequiredMixin, CreateView):
    model = CargaUrea
    form_class = CargaUreaForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargaurea-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Registrar Carga de Urea'
        return context

class CargaUreaUpdateView(AdminRequiredMixin, UpdateView):
    model = CargaUrea
    form_class = CargaUreaForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargaurea-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f"Edit Urea Load from {self.object.fecha.strftime('%d/%m/%Y')}"
        return context

class CargaUreaDeleteView(AdminRequiredMixin, DeleteView):
    model = CargaUrea
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('cargaurea-list')

# --- Supply Purchase CRUD ---
class CompraSuministroListView(AdminRequiredMixin, ListView):
    model = CompraSuministro
    template_name = 'generic_list.html'
    paginate_by = 25

    # --- MÉTODO MODIFICADO ---
    def get_queryset(self):
        queryset = super().get_queryset().order_by('-fecha_compra')
        
        # Lógica de filtrado añadida
        tipo_suministro = self.request.GET.get('tipo_suministro')
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        if tipo_suministro:
            queryset = queryset.filter(tipo_suministro=tipo_suministro)
        if start_date_str and end_date_str:
            # Filtramos por 'fecha_compra__date' para ignorar la hora
            queryset = queryset.filter(fecha_compra__date__range=[date.fromisoformat(start_date_str), date.fromisoformat(end_date_str)])
            
        return queryset

    # --- MÉTODO MODIFICADO ---
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Compras de Suministros', 
            'url_crear': 'comprasuministro-create', 
            'headers': ['#', 'Fecha', 'Proveedor', 'Suministro', 'Precio Total'], 
            'url_update_name': 'comprasuministro-update', 
            'url_delete_name': 'comprasuministro-delete',
            # 'url_exportar_excel': 'comprasuministro-export-excel', # <--- Eliminamos esta línea
        })
        
        # Pasamos los valores de los filtros a la plantilla
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        context['selected_tipo_suministro'] = self.request.GET.get('tipo_suministro', '')
        
        # Pasamos las opciones del modelo para el dropdown del filtro
        context['tipos_suministro'] = CompraSuministro.TIPO_SUMINISTRO_CHOICES

        # ========= INICIO DEL CÓDIGO A AGREGAR =========
        # Habilita el botón de exportación en la plantilla
        context['show_general_export_button'] = True
        # Define la URL a la que apuntará el botón
        context['export_url_name'] = 'comprasuministro-export-excel'
        # ========= FIN DEL CÓDIGO A AGREGAR =========
        
        return context

class CompraSuministroCreateView(AdminRequiredMixin, CreateView):
    model = CompraSuministro
    form_class = CompraSuministroForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('comprasuministro-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Register Supply Purchase'
        return context

class CompraSuministroUpdateView(AdminRequiredMixin, UpdateView):
    model = CompraSuministro
    form_class = CompraSuministroForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('comprasuministro-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Edit Supply Purchase'
        return context

class CompraSuministroDeleteView(AdminRequiredMixin, DeleteView):
    model = CompraSuministro
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('comprasuministro-list')

# --- Checklist Inspection CRUD ---
class ChecklistListView(AdminRequiredMixin, ListView):
    model = ChecklistInspeccion
    template_name = 'generic_list.html'
    paginate_by = 25
    
    # --- MÉTODO MODIFICADO ---
    def get_queryset(self):
        # La lógica de filtrado que ya tenías era correcta, la mantenemos.
        queryset = ChecklistInspeccion.objects.select_related('unidad', 'operador', 'tecnico').order_by('-fecha')
        unidad_id = self.request.GET.get('unidad')
        tecnico_id = self.request.GET.get('tecnico')
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        
        if unidad_id: queryset = queryset.filter(unidad_id=unidad_id)
        if tecnico_id: queryset = queryset.filter(tecnico_id=tecnico_id)
        if start_date_str and end_date_str:
            queryset = queryset.filter(fecha__date__range=[date.fromisoformat(start_date_str), date.fromisoformat(end_date_str)])
            
        return queryset

    # --- MÉTODO MODIFICADO ---
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Registros de Checklist', 
            'url_crear': 'checklist-create', 
            'headers': ['#', 'Unidad', 'Operador', 'Técnico', 'Fecha'], 
            'url_detail_name': 'checklist-detail', 
            'url_update_name': 'checklist-update', 
            'url_delete_name': 'checklist-delete'
        })
        # Pasamos los valores de los filtros a la plantilla
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        if selected_unidad_id := self.request.GET.get('unidad'):
            context['selected_unidad'] = Unidad.objects.filter(pk=selected_unidad_id).first()
        if selected_tecnico_id := self.request.GET.get('tecnico'):
            context['selected_tecnico'] = User.objects.filter(pk=selected_tecnico_id).first()
        return context

class ChecklistDetailView(AdminRequiredMixin, DetailView):
    model = ChecklistInspeccion
    template_name = 'checklist_detail.html'

    # --- MÉTODO AÑADIDO/MODIFICADO ---
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        checklist = self.get_object()
        
        items_procesados = []
        # Iteramos sobre los campos en la vista, donde sí está permitido
        for field in ChecklistInspeccion._meta.get_fields():
            # Nos aseguramos de procesar solo los campos de estado (que tienen opciones)
            if hasattr(field, 'choices') and field.choices:
                
                # Obtenemos el nombre del campo de observación correspondiente
                obs_field_name = f"{field.name}_obs"
                
                # === INICIO LÓGICA S3 ===
                # Obtenemos el campo de foto
                foto_field_name = f"{field.name}_foto"
                foto_obj = getattr(checklist, foto_field_name, None)
                # === FIN LÓGICA S3 ===

                # Creamos un diccionario con la información limpia para el template
                items_procesados.append({
                    'componente': field.verbose_name.title(),
                    'estado': getattr(checklist, field.name, ''),
                    'observacion': getattr(checklist, obs_field_name, "Sin observación."),
                    'foto': foto_obj # Pasamos el objeto de foto (o None)
                })

        # Agregamos la lista ya procesada al contexto
        context['items_procesados'] = items_procesados
        return context

class ChecklistCreateView(AdminRequiredMixin, CreateView):
    model = ChecklistInspeccion
    form_class = ChecklistInspeccionForm
    template_name = 'checklist_form.html'
    success_url = reverse_lazy('checklist-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = "Create New Checklist Record"
        form = context['form']
        field_groups = {
            'Estructura Exterior': ['cristales', 'espejos', 'logos', 'num_economico', 'puertas', 'cofre', 'parrilla', 'defensas', 'faros', 'plafoneria', 'stops', 'direccionales', 'tapiceria', 'instrumentos', 'carroceria', 'piso', 'costados', 'escape', 'pintura', 'franjas', 'loderas', 'extintor', 'senalamientos', 'estado_general'],
            'Mecánica y Motor': ['motor', 'caja', 'diferenciales', 'suspension_delantera', 'suspension_trasera', 'fugas_combustible', 'fugas_aceite', 'estado_llantas', 'presion_llantas', 'purga_tanques', 'estado_balatas', 'amortiguadores_delanteros', 'amortiguadores_traseros', 'rines_aluminio', 'mangueras_servicio', 'tarjeta_llave', 'revision_fusibles', 'revision_luces','revision_fuga_aire']
        }
        structured_form = {}
        for group_name, field_list in field_groups.items():
            structured_form[group_name] = []
            for field_name in field_list:
                if field_name in form.fields:
                    obs_field_name = f"{field_name}_obs"
                    # --- INICIO DE LÓGICA MODIFICADA ---
                    foto_field_name = f"{field_name}_foto"
                    structured_form[group_name].append({
                        'status': form[field_name], 
                        'observation': form[obs_field_name] if obs_field_name in form.fields else None,
                        'evidence': form[foto_field_name] if foto_field_name in form.fields else None
                    })
                    # --- FIN DE LÓGICA MODIFICADA ---
        context['structured_form'] = structured_form
        return context

    # === INICIO: MÉTODO MODIFICADO PARA S3 MANUAL ===
    def form_valid(self, form):
        # Asigna el tecnico ANTES de save(commit=False)
        form.instance.tecnico = self.request.user 
        try:
            checklist = form.save(commit=False)
            unidad = checklist.unidad 
            
            # La fecha se asigna por 'default=timezone.now' en el modelo
            # o por el valor del formulario.
            fecha_str = checklist.fecha.strftime('%Y-%m-%d')
            
            # Iterar y subir archivos
            for field_name, archivo in self.request.FILES.items():
                _nombre_base, extension = os.path.splitext(archivo.name)
                # Define una ruta única en S3
                s3_path = f"flota/checklists/{unidad.nombre}/{fecha_str}/{field_name}{extension}"
                ruta_guardada = _subir_archivo_a_s3(archivo, s3_path)
                
                if ruta_guardada:
                    setattr(checklist, field_name, ruta_guardada)
                else:
                    messages.error(self.request, f"Error al subir el archivo {field_name}.")
                    return self.form_invalid(form)

            # Guardar el objeto en la DB
            checklist.save() 
            
            # Asignar el objeto guardado a self.object para la redirección
            self.object = checklist 
            
            messages.success(self.request, "Checklist successfully created.")
            # Usar get_success_url() para la redirección
            return redirect(self.get_success_url())
        
        except Exception as e:
            messages.error(self.request, f"Ocurrió un error al guardar: {e}")
            return self.form_invalid(form)
    # === FIN: MÉTODO MODIFICADO ===

class ChecklistUpdateView(AdminRequiredMixin, UpdateView):
    model = ChecklistInspeccion
    form_class = ChecklistInspeccionForm
    template_name = 'checklist_form.html'
    success_url = reverse_lazy('checklist-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f"Edit Checklist for {self.object.unidad}"
        form = context['form']
        field_groups = {
            'Estructura Exterior': ['cristales', 'espejos', 'logos', 'num_economico', 'puertas', 'cofre', 'parrilla', 'defensas', 'faros', 'plafoneria', 'stops', 'direccionales', 'tapiceria', 'instrumentos', 'carroceria', 'piso', 'costados', 'escape', 'pintura', 'franjas', 'loderas', 'extintor', 'senalamientos', 'estado_general'],
            'Mecánica y Motor': ['motor', 'caja', 'diferenciales', 'suspension_delantera', 'suspension_trasera', 'fugas_combustible', 'fugas_aceite', 'estado_llantas', 'presion_llantas', 'purga_tanques', 'estado_balatas', 'amortiguadores_delanteros', 'amortiguadores_traseros', 'rines_aluminio', 'mangueras_servicio', 'tarjeta_llave', 'revision_fusibles', 'revision_luces', 'revision_fuga_aire']
        }
        structured_form = {}
        for group_name, field_list in field_groups.items():
            structured_form[group_name] = []
            for field_name in field_list:
                if field_name in form.fields:
                    obs_field_name = f"{field_name}_obs"
                    # --- INICIO DE LÓGICA MODIFICADA ---
                    foto_field_name = f"{field_name}_foto"
                    structured_form[group_name].append({
                        'status': form[field_name], 
                        'observation': form[obs_field_name] if obs_field_name in form.fields else None,
                        'evidence': form[foto_field_name] if foto_field_name in form.fields else None
                    })
                    # --- FIN DE LÓGICA MODIFICADA ---
        context['structured_form'] = structured_form
        return context

    # === INICIO: MÉTODO MODIFICADO PARA S3 MANUAL ===
    def form_valid(self, form):
        try:
            checklist_original = self.get_object()
            checklist = form.save(commit=False)
            unidad = checklist.unidad
            fecha_str = checklist.fecha.strftime('%Y-%m-%d')

            # Iterar y subir/actualizar archivos
            for field_name, archivo in self.request.FILES.items():
                # 1. Eliminar el archivo antiguo
                ruta_antigua_field = getattr(checklist_original, field_name)
                if ruta_antigua_field and hasattr(ruta_antigua_field, 'name'):
                    _eliminar_archivo_de_s3(ruta_antigua_field.name)
                    
                # 2. Subir el archivo nuevo
                _nombre_base, extension = os.path.splitext(archivo.name)
                s3_path = f"flota/checklists/{unidad.nombre}/{fecha_str}/{field_name}{extension}"
                ruta_guardada = _subir_archivo_a_s3(archivo, s3_path)
                
                if ruta_guardada:
                    setattr(checklist, field_name, ruta_guardada)
                else:
                    messages.error(self.request, f"Error al actualizar el archivo {field_name}.")
                    return self.form_invalid(form)

            # Guardar el objeto en la DB
            checklist.save()
            
            # Guardar relaciones ManyToMany si las hubiera
            form.save_m2m() 
            
            self.object = checklist
            messages.success(self.request, "Checklist successfully updated.")
            return redirect(self.get_success_url())

        except Exception as e:
            messages.error(self.request, f"Ocurrió un error al actualizar: {e}")
            return self.form_invalid(form)
    # === FIN: MÉTODO MODIFICADO ===

class ChecklistDeleteView(AdminRequiredMixin, DeleteView):
    model = ChecklistInspeccion
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('checklist-list')

    # === INICIO: MÉTODO AÑADIDO PARA BORRADO MANUAL EN S3 ===
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        try:
            # Iterar sobre todos los campos del modelo
            for field in ChecklistInspeccion._meta.get_fields():
                # Buscar campos de archivo (asumiendo que terminan en _foto)
                if field.name.endswith('_foto'):
                    file_field = getattr(self.object, field.name)
                    # file_field.name contiene la ruta completa 'media/flota/...'
                    if file_field and hasattr(file_field, 'name'):
                        _eliminar_archivo_de_s3(file_field.name)
            
            # Llamar al método de borrado original de la base de datos
            response = super().delete(request, *args, **kwargs)
            
            messages.success(request, f"Checklist para '{self.object.unidad}' eliminado exitosamente.")
            return response
            
        except Exception as e:
            messages.error(request, f"Ocurrió un error al eliminar el checklist y sus archivos: {e}")
            return redirect(self.get_success_url())
    # === FIN: MÉTODO AÑADIDO ===


# --- Tire Inspection CRUD ---
class LlantasInspeccionListView(AdminRequiredMixin, ListView):
    model = LlantasInspeccion
    template_name = 'generic_list.html'
    paginate_by = 25

    # --- MÉTODO MODIFICADO ---
    def get_queryset(self):
        queryset = LlantasInspeccion.objects.select_related('unidad', 'tecnico').order_by('-fecha')
        
        # Lógica de filtrado añadida
        unidad_id = self.request.GET.get('unidad')
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        if unidad_id:
            queryset = queryset.filter(unidad_id=unidad_id)
        if start_date_str and end_date_str:
            queryset = queryset.filter(fecha__date__range=[date.fromisoformat(start_date_str), date.fromisoformat(end_date_str)])
            
        return queryset

    # --- MÉTODO MODIFICADO ---
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Inspecciones de Llantas', 
            'url_crear': 'llantas-create', 
            'headers': ['#', 'Unidad', 'Kilometraje', 'Técnico', 'Fecha'], 
            'url_detail_name': 'llantas-detail', 
            'url_update_name': 'llantas-update', 
            'url_delete_name': 'llantas-delete'
        })
        # Pasamos los valores de los filtros a la plantilla
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        if selected_unidad_id := self.request.GET.get('unidad'):
            context['selected_unidad'] = Unidad.objects.filter(pk=selected_unidad_id).first()
        
        # ========= INICIO DEL CÓDIGO A AGREGAR =========
        context['show_general_export_button'] = True
        context['export_url_name'] = 'llantas-export-excel'
        # ========= FIN DEL CÓDIGO A AGREGAR =========
        return context

class LlantasInspeccionDetailView(AdminRequiredMixin, DetailView):
    model = LlantasInspeccion
    template_name = 'llantas_inspeccion_detail.html'

class LlantasInspeccionCreateView(AdminRequiredMixin, CreateView):
    model = LlantasInspeccion
    form_class = LlantasInspeccionForm
    template_name = 'llantas_form_unificado.html'
    success_url = reverse_lazy('llantas-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = "Create Tire Inspection"
        FormSet = inlineformset_factory(LlantasInspeccion, LlantaDetalle, form=LlantaDetalleForm, extra=6, max_num=6, can_delete=False)
        if self.request.POST:
            context['formset'] = FormSet(self.request.POST, prefix='llantas')
        else:
            context['formset'] = FormSet(prefix='llantas', initial=[{'posicion': f'Posición {i}'} for i in range(1, 7)])
        context['user_role'] = 'admin'
        return context
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                form.instance.tecnico = self.request.user
                self.object = form.save()
                formset.instance = self.object
                formset.save()
            messages.success(self.request, "Tire inspection created.")
            return redirect(self.success_url)
        return self.form_invalid(form)

class LlantasInspeccionUpdateView(AdminRequiredMixin, UpdateView):
    model = LlantasInspeccion
    # --- CAMBIO 1: El formulario principal ahora se define aquí ---
    form_class = LlantasInspeccionForm 
    template_name = 'llantas_form_unificado.html'
    success_url = reverse_lazy('llantas-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # --- INICIO: CÓDIGO AÑADIDO PARA CORREGIR DATOS EN BLANCO ---
        # Obtenemos la inspección que se está editando
        inspeccion = self.get_object()
        
        # Pasamos los datos que faltaban a la plantilla
        context['titulo'] = f"Editar Inspección de Llantas: {inspeccion.unidad.nombre}"
        context['unidad'] = inspeccion.unidad
        context['fecha_actual'] = inspeccion.fecha
        context['form_mode'] = 'update' # Variable para controlar los botones

        # Creamos y pasamos el formulario de KM con los datos existentes
        context['km_form'] = LlantasKmForm(
            initial={'km': inspeccion.km},
            unidad=inspeccion.unidad
        )
        # --- FIN: CÓDIGO AÑADIDO ---
        
        FormSet = inlineformset_factory(LlantasInspeccion, LlantaDetalle, form=LlantaDetalleForm, extra=0, max_num=6, can_delete=True)
        
        if self.request.POST:
            context['formset'] = FormSet(self.request.POST, instance=self.object, prefix='llantas')
        else:
            context['formset'] = FormSet(instance=self.object, prefix='llantas')
            
        return context

    def post(self, request, *args, **kwargs):
        """Maneja el envío del formulario principal y el formset de detalles."""
        self.object = self.get_object()
        
        # Validamos ambos formularios: el de KM y el de los detalles de llantas
        form = self.get_form()
        km_form = LlantasKmForm(request.POST, unidad=self.object.unidad)
        
        FormSet = inlineformset_factory(LlantasInspeccion, LlantaDetalle, form=LlantaDetalleForm, extra=0)
        formset = FormSet(request.POST, instance=self.object, prefix='llantas')

        if form.is_valid() and km_form.is_valid() and formset.is_valid():
            return self.form_valid(form, km_form, formset)
        else:
            # Si hay un error, volvemos a renderizar todo con los errores
            return self.form_invalid(form, km_form, formset)

    def form_valid(self, form, km_form, formset):
        """Si todo es válido, guarda los cambios."""
        with transaction.atomic():
            # Actualiza el KM de la inspección desde el km_form
            inspeccion = form.save(commit=False)
            inspeccion.km = km_form.cleaned_data['km']
            inspeccion.save()
            
            # Guarda los cambios en los detalles de las llantas
            formset.save()

        messages.success(self.request, "Inspección de llantas actualizada correctamente.")
        return redirect(self.success_url)

    def form_invalid(self, form, km_form, formset):
        """Si hay errores, vuelve a mostrar el formulario con los datos y errores."""
        messages.error(self.request, "Por favor, corrija los errores marcados.")
        context = self.get_context_data()
        context['form'] = form # El form principal (vacío en este caso, pero necesario)
        context['km_form'] = km_form # El form de KM con sus errores
        context['formset'] = formset # El formset de llantas con sus errores
        return self.render_to_response(context)

class LlantasInspeccionDeleteView(AdminRequiredMixin, DeleteView):
    model = LlantasInspeccion
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('llantas-list')


# ===================================================================
# 3. VIEWS FOR THE NEW DIVIDED WORKFLOW
# ===================================================================

# --- STAGE 1: TECHNICIAN or MANAGER ---

class SeleccionarUnidadView(IniciaProcesoRequiredMixin, ListView):
    """Paso 1 (Común): Elegir una unidad para iniciar el proceso."""
    model = Unidad
    template_name = 'tecnico_seleccionar_unidad.html'
    context_object_name = 'unidades'

    def dispatch(self, request, *args, **kwargs):
        """
        Intercepta la solicitud ANTES de que la vista se ejecute.
        Comprueba si el usuario tiene un checklist guardado en la sesión que
        aún no ha sido vinculado a una inspección de llantas.
        """
        checklist_id_pendiente = request.session.get('proceso_checklist_id')

        if checklist_id_pendiente:
            try:
                # Verificamos que el checklist realmente exista en la BD
                checklist = ChecklistInspeccion.objects.get(pk=checklist_id_pendiente)
                unidad_pk = checklist.unidad.id
                
                # Enviamos un mensaje de advertencia al usuario
                messages.warning(
                    request, 
                    f"Tiene una inspección de llantas pendiente para la unidad '{checklist.unidad.nombre}'. "
                    "Por favor, complétela para poder iniciar un nuevo proceso."
                )
                # Lo redirigimos a la pantalla de llenado de llantas
                return redirect('proceso-llantas', unidad_pk=unidad_pk)

            except ChecklistInspeccion.DoesNotExist:
                # Si el checklist no existe (por alguna razón), limpiamos la sesión
                request.session.pop('proceso_checklist_id', None)
        
        # Si no hay nada pendiente, la vista continúa de forma normal
        return super().dispatch(request, *args, **kwargs)

    # --- INICIO DE CÓDIGO MODIFICADO ---
    def get_queryset(self):
        """
        Filtra las unidades según el parámetro de búsqueda 'q' en la URL.
        """
        queryset = super().get_queryset().order_by('nombre')
        search_query = self.request.GET.get('q')
        if search_query:
            # Filtra por el nombre de la unidad, sin ser sensible a mayúsculas/minúsculas
            queryset = queryset.filter(nombre__icontains=search_query)
        return queryset

    def get_context_data(self, **kwargs):
        """
        Añade el término de búsqueda al contexto para que pueda ser mostrado
        en el campo de texto del formulario después de una búsqueda.
        """
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        return context

class ProcesoChecklistView(IniciaProcesoRequiredMixin, FormView):
    """Step 2 (Common): Fill out the Checklist."""
    form_class = ChecklistInspeccionForm
    template_name = 'checklist_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        context['titulo'] = f"Step 1: Checklist for Unit {unidad.nombre}"
        context['unidad'] = unidad
        form = context.get('form')
        field_groups = {
            'Estructura Exterior': ['cristales', 'espejos', 'logos', 'num_economico', 'puertas', 'cofre', 'parrilla', 'defensas', 'faros', 'plafoneria', 'stops', 'direccionales', 'tapiceria', 'instrumentos', 'carroceria', 'piso', 'costados', 'escape', 'pintura', 'franjas', 'loderas', 'extintor', 'senalamientos', 'estado_general'],
            'Mecánica y Motor': ['motor', 'caja', 'diferenciales', 'suspension_delantera', 'suspension_trasera', 'fugas_combustible', 'fugas_aceite', 'estado_llantas', 'presion_llantas', 'purga_tanques', 'estado_balatas', 'amortiguadores_delanteros', 'amortiguadores_traseros', 'rines_aluminio', 'mangueras_servicio', 'tarjeta_llave', 'revision_fusibles', 'revision_luces','revision_fuga_aire']
        }
        structured_form = {}
        for group_name, field_list in field_groups.items():
            structured_form[group_name] = []
            for field_name in field_list:
                if field_name in form.fields:
                    obs_field_name = f"{field_name}_obs"
                    # --- INICIO DE LÓGICA MODIFICADA ---
                    foto_field_name = f"{field_name}_foto"
                    structured_form[group_name].append({
                        'status': form[field_name], 
                        'observation': form[obs_field_name] if obs_field_name in form.fields else None,
                        'evidence': form[foto_field_name] if foto_field_name in form.fields else None
                    })
                    # --- FIN DE LÓGICA MODIFICADA ---
        context['structured_form'] = structured_form
        return context

    def get_initial(self):
        """
        --- MÉTODO ACTUALIZADO ---
        Establece valores iniciales. Si el usuario es un 'Encargado',
        pre-llena todos los campos de estado como 'BIEN'.
        """
        # Primero, limpiamos la sesión de datos de procesos anteriores.
        self.request.session.pop('proceso_checklist_id', None)
        
        unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        initial = {'unidad': unidad}

        # Verificamos si el usuario pertenece al grupo 'Encargado'.
        if es_encargado(self.request.user):
            # Iteramos sobre todos los campos definidos en el modelo del checklist.
            for field in ChecklistInspeccion._meta.get_fields():
                # Buscamos solo los campos que son de tipo CharField y tienen opciones (BIEN/MALO).
                if isinstance(field, models.CharField) and hasattr(field, 'choices') and field.choices:
                    # Añadimos el campo al diccionario de valores iniciales con el valor 'BIEN'.
                    initial[field.name] = 'BIEN'
        
        return initial

    # === INICIO: MÉTODO MODIFICADO PARA S3 MANUAL ===
    def form_valid(self, form):
        """
        Guarda el formulario manualmente, subiendo archivos a S3
        y manejando la transacción de forma atómica.
        """
        try:
            with transaction.atomic():
                # 1. Asigna los datos que no vienen del formulario
                form.instance.tecnico = self.request.user
                form.instance.unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
                
                # 2. Crea el objeto en memoria (aún no en la BD)
                checklist_obj = form.save(commit=False)
                
                # 3. Iterar y subir archivos
                
                # === INICIO DE LA CORRECCIÓN ===
                # Obtenemos la fecha actual AHORA.
                # El campo `checklist_obj.fecha` es 'None' en este punto
                # porque `auto_now_add=True` solo lo asigna cuando se
                # ejecuta el .save() en la base de datos.
                # Usamos timezone.now() para la fecha.
                fecha_actual = timezone.now()
                fecha_str = fecha_actual.strftime('%Y-%m-%d')
                # === FIN DE LA CORRECCIÓN ===
                
                for field_name, archivo in self.request.FILES.items():
                    _nombre_base, extension = os.path.splitext(archivo.name)
                    # Define la ruta de S3 usando la fecha_str que acabamos de crear
                    s3_path = f"flota/checklists/{checklist_obj.unidad.nombre}/{fecha_str}/{field_name}{extension}"
                    
                    ruta_guardada = _subir_archivo_a_s3(archivo, s3_path)
                    
                    if ruta_guardada:
                        # Asigna la ruta de S3 al campo del modelo
                        setattr(checklist_obj, field_name, ruta_guardada)
                    else:
                        messages.error(self.request, f"Error al subir el archivo {field_name}.")
                        # Esto abortará la transacción atómica
                        raise Exception(f"Fallo al subir {field_name}")
                
                # 4. Ahora sí, guardar el objeto en la base de datos
                # En este punto, `auto_now_add=True` asignará la fecha
                # a checklist_obj.fecha. Será casi idéntica a la que
                # usamos para la ruta de S3.
                checklist_obj.save()
            
            # 5. Si todo salió bien, guarda el ID en la sesión
            self.request.session['proceso_checklist_id'] = checklist_obj.id
            messages.success(self.request, "Checklist guardado. Ahora, por favor, complete la inspección de llantas.")
            
            # 6. Redirige al siguiente paso del proceso.
            return redirect('proceso-llantas', unidad_pk=self.kwargs['unidad_pk'])

        except Exception as e:
            # Captura el error de la subida de S3 o cualquier otro.
            messages.error(self.request, f"Ocurrió un error al guardar el checklist: {e}")
            return self.form_invalid(form)
    # === FIN: MÉTODO MODIFICADO ===
        
class ProcesoLlantasView(IniciaProcesoRequiredMixin, TemplateView):
    """Step 3 (Common): Fill out Tire form and send to PENDING."""
    template_name = 'llantas_form_unificado.html'

    def dispatch(self, request, *args, **kwargs):
        if 'proceso_checklist_id' not in request.session:
            messages.error(request, 'Error: You must complete the checklist before continuing.')
            return redirect('tecnico-seleccionar-unidad')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        LlantaFormSet = formset_factory(LlantaDetalleForm, extra=6, max_num=6)
        posiciones_llantas = [{'posicion': f'Posición {i}'} for i in range(1, 7)]
        context['titulo'] = f"Step 2: Tire Format for {unidad.nombre}"
        context['unidad'] = unidad
        context['km_form'] = LlantasKmForm(initial={'km': unidad.km_actual}, unidad=unidad)
        context['formset'] = LlantaFormSet(initial=posiciones_llantas, prefix='llantas')
        context['fecha_actual'] = date.today()
        return context

    def post(self, request, *args, **kwargs):
        unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        LlantaFormSet = formset_factory(LlantaDetalleForm, extra=6, max_num=6)
        
        km_form = LlantasKmForm(request.POST, unidad=unidad)
        formset = LlantaFormSet(request.POST, prefix='llantas')

        if not km_form.is_valid() or not formset.is_valid():
            messages.error(request, 'Por favor, corrija los errores marcados en rojo.')
            
            # --- INICIO: CÓDIGO DE DIAGNÓSTICO ---
            # Estas líneas imprimirán información en tu consola de runserver.
            print("=" * 60)
            print("DIAGNÓSTICO DE FORMULARIO DE LLANTAS (POST INVÁLIDO)")
            print(f"Timestamp: {datetime.now()}")
            print(f"KM Form es válido: {km_form.is_valid()}")
            if not km_form.is_valid():
                print(f" -> Errores en KM Form: {km_form.errors.as_json()}")

            print(f"Formset es válido: {formset.is_valid()}")
            if not formset.is_valid():
                print(f" -> Errores en Formset: {formset.errors}")

            # Imprime los datos crudos que recibió el formset del POST.
            # Esto nos mostrará si la información de las llantas está llegando al servidor.
            print("\n--- DATOS CRUDOS RECIBIDOS POR EL FORMSET ---")
            for i, form in enumerate(formset.forms):
                # form.data contiene todos los datos del POST
                # form.prefix nos da el prefijo correcto para este formulario (ej. 'llantas-0')
                print(f"  Formulario {i} ({form.prefix}):")
                print(f"    mm: '{form.data.get(f'{form.prefix}-mm')}'")
                print(f"    marca: '{form.data.get(f'{form.prefix}-marca')}'")
                print(f"    modelo: '{form.data.get(f'{form.prefix}-modelo')}'")
                print(f"    medida: '{form.data.get(f'{form.prefix}-medida')}'")
                print(f"    presion: '{form.data.get(f'{form.prefix}-presion')}'")
            print("=" * 60)
            # --- FIN: CÓDIGO DE DIAGNÓSTICO ---

            context = {
                'titulo': f"Step 2: Tire Format for {unidad.nombre}",
                'unidad': unidad,
                'km_form': km_form,
                'formset': formset,
                'fecha_actual': date.today(),
            }
            return self.render_to_response(context)

        # Si todo es válido, el proceso continúa para guardar...
        # ... (el resto del código de guardado se queda exactamente igual) ...
        checklist_id = request.session.get('proceso_checklist_id')
        if not checklist_id:
            messages.error(request, "La sesión ha expirado. Por favor, inicie de nuevo.")
            return redirect('tecnico-seleccionar-unidad')

        try:
            with transaction.atomic():
                checklist_obj = get_object_or_404(ChecklistInspeccion, pk=checklist_id)
                km_llantas = km_form.cleaned_data['km']
                llantas_inspeccion_obj = LlantasInspeccion.objects.create(
                    unidad=unidad, tecnico=request.user, km=km_llantas
                )
                for form_data in formset.cleaned_data:
                    if form_data and form_data.get('mm'):
                        LlantaDetalle.objects.create(inspeccion=llantas_inspeccion_obj, **form_data)
                
                # Aquí se crea el ProcesoCarga
                proceso_carga_obj = ProcesoCarga.objects.create(
                    unidad=unidad, checklist=checklist_obj, inspeccion_llantas=llantas_inspeccion_obj, 
                    tecnico_inicia=request.user, status='PENDIENTE'
                )
                
                # ========= INICIO DE LÓGICA AÑADIDA (EN_PROCESO) =========
                try:
                    asignacion_del_dia = AsignacionRevision.objects.get(
                        unidad=proceso_carga_obj.unidad,
                        # USA LA HORA LOCAL PARA OBTENER LA FECHA
                        fecha_revision=timezone.localdate(proceso_carga_obj.fecha_inicio),
                        status='PENDIENTE' 
                    )
                    asignacion_del_dia.status = 'EN_PROCESO'
                    asignacion_del_dia.save()
                except AsignacionRevision.DoesNotExist:
                    pass 
                # ========= FIN DE LA LÓGICA AÑADIDA =========
                
                if km_llantas > unidad.km_actual:
                    unidad.km_actual = km_llantas
                    unidad.save()

            request.session.pop('proceso_checklist_id', None)
            messages.success(request, f"Proceso para {unidad.nombre} enviado a pendientes.")
            
            if es_encargado(request.user):
                return redirect('encargado-pendientes-list')
            return redirect('tecnico-seleccionar-unidad')
        
        except Exception as e:
            messages.error(request, f"Ocurrió un error inesperado al guardar: {e}.")
            context = {'km_form': km_form, 'formset': formset}
            return self.render_to_response(context)

class EncargadoPendientesListView(EncargadoRequiredMixin, ListView):
    """Página principal del Encargado: Muestra unidades pendientes."""
    model = ProcesoCarga
    template_name = 'encargado_pendientes_list.html'
    context_object_name = 'procesos_pendientes'

    def get_queryset(self):
        return ProcesoCarga.objects.filter(status='PENDIENTE').select_related('unidad', 'tecnico_inicia')

    # --- MÉTODO AÑADIDO ---
    def dispatch(self, request, *args, **kwargs):
        """
        Intercepta al Encargado para verificar si dejó una carga de urea a medias.
        Busca en la sesión cualquier clave que empiece con 'diesel_data_proceso_'.
        """
        # Iteramos sobre una copia de las claves de la sesión para poder modificarla
        for key in list(request.session.keys()):
            if key.startswith('diesel_data_proceso_'):
                # Extraemos el ID del proceso de la clave de la sesión
                proceso_pk = key.split('_')[-1]
                try:
                    proceso = ProcesoCarga.objects.get(pk=proceso_pk, status='PENDIENTE')
                    messages.warning(
                        request,
                        f"Tiene pendiente finalizar el proceso para la unidad '{proceso.unidad.nombre}'. "
                        "Por favor, complete la carga de urea."
                    )
                    # Redirigimos directamente al último paso que le falta
                    return redirect('encargado-proceso-urea', proceso_pk=proceso.pk)
                
                except ProcesoCarga.DoesNotExist:
                    # El proceso ya no existe o fue completado. Limpiamos la sesión.
                    request.session.pop(key, None)

        # Si no hay nada pendiente, la vista continúa de forma normal
        return super().dispatch(request, *args, **kwargs)

class EncargadoProcesoDieselView(EncargadoRequiredMixin, FormView):
    """Manager fills in the Diesel data."""
    form_class = CargaDieselForm # Usa nuestro formulario inteligente
    template_name = 'generic_form.html'

    def get_proceso(self):
        return get_object_or_404(ProcesoCarga, pk=self.kwargs['proceso_pk'], status='PENDIENTE')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        proceso = self.get_proceso()
        context['titulo'] = f"Step 3: Diesel Load for {proceso.unidad.nombre}"
        context['url_cancelar'] = reverse('encargado-pendientes-list')
        return context

    def get_form_kwargs(self):
        """
        Esta es la función clave. Pasa la instancia de la unidad al formulario.
        """
        kwargs = super().get_form_kwargs()
        proceso = self.get_proceso()
        # Aquí se inyecta la unidad, lo que permite la lógica en el __init__ del form.
        kwargs['unidad'] = proceso.unidad
        return kwargs
    
    def get_initial(self):
        proceso = self.get_proceso()
        ultima_carga = CargaDiesel.objects.filter(unidad=proceso.unidad).order_by('-fecha').first()
        return {
            'unidad': proceso.unidad,
            'operador': proceso.checklist.operador,
            'km_actual': proceso.inspeccion_llantas.km,
            'persona_relleno': self.request.user.get_full_name() or self.request.user.username,
            'cinchos_anteriores': ultima_carga.cinchos_actuales if ultima_carga else "",
        }

    def form_valid(self, form):
        diesel_data = form.cleaned_data
        proceso = self.get_proceso() # Necesario para la ruta S3
        fecha_str = timezone.now().strftime('%Y-%m-%d')

        # --- INICIO: Lógica de subida manual de archivos ---
        try:
            foto_motor_file = self.request.FILES.get('foto_motor')
            foto_thermo_file = self.request.FILES.get('foto_thermo')

            if foto_motor_file:
                _nombre_base, extension = os.path.splitext(foto_motor_file.name)
                s3_path = f"flota/cargas_diesel/{proceso.unidad.nombre}/{fecha_str}/motor{extension}"
                ruta_guardada = _subir_archivo_a_s3(foto_motor_file, s3_path)
                if ruta_guardada:
                    diesel_data['foto_motor'] = ruta_guardada
                else:
                    messages.error(self.request, "Error al subir la foto del motor.")
                    return self.form_invalid(form)

            if foto_thermo_file:
                _nombre_base, extension = os.path.splitext(foto_thermo_file.name)
                s3_path = f"flota/cargas_diesel/{proceso.unidad.nombre}/{fecha_str}/thermo{extension}"
                ruta_guardada = _subir_archivo_a_s3(foto_thermo_file, s3_path)
                if ruta_guardada:
                    diesel_data['foto_thermo'] = ruta_guardada
                else:
                    messages.error(self.request, "Error al subir la foto del thermo.")
                    return self.form_invalid(form)
        
        except Exception as e:
            messages.error(self.request, f"Ocurrió un error al procesar las imágenes: {e}")
            return self.form_invalid(form)
        # --- FIN: Lógica de subida manual de archivos ---

        diesel_data['operador_id'] = diesel_data.pop('operador').id
        if 'unidad' in diesel_data: del diesel_data['unidad']
        
        for key, value in diesel_data.items():
            if isinstance(value, Decimal):
                diesel_data[key] = str(value)
            # --- AÑADIDO: Asegurar que los campos ImageField (None) no den error ---
            elif isinstance(value, models.fields.files.ImageFieldFile):
                if not value:
                    diesel_data[key] = None

        self.request.session[f'diesel_data_proceso_{self.kwargs["proceso_pk"]}'] = diesel_data
        return redirect('encargado-proceso-urea', proceso_pk=self.kwargs['proceso_pk'])

class EncargadoProcesoUreaView(EncargadoRequiredMixin, FormView):
    """Manager fills in Urea and completes the process."""
    form_class = CargaUreaForm
    template_name = 'generic_form.html'

    def get_proceso(self):
        return get_object_or_404(ProcesoCarga, pk=self.kwargs['proceso_pk'], status='PENDIENTE')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        proceso = self.get_proceso()
        context['titulo'] = f"Paso 4: Carga de Urea y Finalizar para {proceso.unidad.nombre}"
        context['url_cancelar'] = reverse('encargado-pendientes-list')
        return context
    
    def post(self, request, *args, **kwargs):
        """
        Interceptamos el POST para imprimir y MOSTRAR los errores si el formulario es inválido.
        """
        form = self.get_form()
        if form.is_valid():
            return self.form_valid(form)
        else:
            # Imprimir los errores en la consola
            print("=====================================================")
            print("==> ERROR: EL FORMULARIO DE UREA ES INVÁLIDO <==")
            print(form.errors.as_json())
            print("=====================================================")

            # Construir un mensaje de error detallado para mostrar en la pantalla
            error_message = "El formulario contiene errores. "
            for field, errors in form.errors.items():
                clean_field = field.replace('_', ' ').capitalize()
                error_list = '; '.join(errors)
                error_message += f"Campo '{clean_field}': {error_list}. "
            
            messages.error(request, error_message)
            
            return self.form_invalid(form)

    def form_valid(self, form):
        proceso = self.get_proceso()
        diesel_data_key = f'diesel_data_proceso_{proceso.pk}'
        diesel_data = self.request.session.get(diesel_data_key)

        if not diesel_data:
            messages.error(self.request, "Error de Sesión: No se encontraron los datos del diésel. Por favor, vuelva a empezar desde el paso anterior.")
            return redirect('encargado-proceso-diesel', proceso_pk=proceso.pk)

        # Convertir valores decimales desde la sesión de forma segura
        for campo in ['lts_diesel', 'lts_thermo', 'hrs_thermo', 'costo']:
            if campo in diesel_data and diesel_data[campo] is not None:
                try:
                    diesel_data[campo] = Decimal(diesel_data[campo])
                except (TypeError, ValueError):
                    messages.error(self.request, f"Error de datos de sesión: El valor para '{campo}' no es válido.")
                    return self.form_invalid(form)

        try:
            # Esta variable nos dirá si necesitamos recalcular urea
            se_agrego_urea = False

            with transaction.atomic():
                operador_id = diesel_data.pop('operador_id', None)
                if not operador_id:
                    raise ValueError("No se encontró el ID del operador en los datos de la sesión.")
                
                operador = get_object_or_404(Operador, pk=operador_id)
                
                # Crear la carga de diésel (ya no dispara el recálculo)
                carga_diesel_obj = CargaDiesel.objects.create(unidad=proceso.unidad, operador=operador, **diesel_data)
                
                carga_urea_obj = None
                litros_urea_cargados = form.cleaned_data.get('litros_cargados')
                
                # Solo procesar si el usuario ingresó un valor numérico mayor a cero
                if litros_urea_cargados and litros_urea_cargados > 0:
                    urea_obj = form.save(commit=False)
                    urea_obj.unidad = proceso.unidad
                    
                    # --- INICIO: LÓGICA DE SUBIDA DE FOTO UREA ---
                    foto_urea_file = self.request.FILES.get('foto_urea')
                    if foto_urea_file:
                        fecha_str = timezone.now().strftime('%Y-%m-%d')
                        _nombre_base, extension = os.path.splitext(foto_urea_file.name)
                        s3_path = f"flota/cargas_urea/{proceso.unidad.nombre}/{fecha_str}/urea{extension}"
                        ruta_guardada = _subir_archivo_a_s3(foto_urea_file, s3_path)
                        
                        if ruta_guardada:
                            urea_obj.foto_urea = ruta_guardada
                        else:
                            messages.error(self.request, "Error al subir la foto de la bomba de urea.")
                            # Abortar la transacción atómica
                            raise Exception("Fallo al subir foto_urea")
                    # --- FIN: LÓGICA DE SUBIDA DE FOTO UREA ---

                    urea_obj.save() # (ya no dispara el recálculo)
                    carga_urea_obj = urea_obj
                    se_agrego_urea = True # Marcamos para recalcular después

                # Actualizar y finalizar el proceso principal
                proceso.carga_diesel = carga_diesel_obj
                proceso.carga_urea = carga_urea_obj
                proceso.encargado_finaliza = self.request.user
                
                proceso.fecha_fin = timezone.now() # Usar timezone.now()
                
                proceso.status = 'COMPLETADO'
                proceso.save()
                
                # Lógica para actualizar la AsignacionRevision
                try:
                    asignacion_del_dia = AsignacionRevision.objects.get(
                        unidad=proceso.unidad,
                        # USA LA HORA LOCAL PARA OBTENER LA FECHA
                        fecha_revision=timezone.localdate(proceso.fecha_inicio),
                        status__in=['PENDIENTE', 'EN_PROCESO'] 
                    )
                    asignacion_del_dia.status = 'TERMINADO'
                    asignacion_del_dia.save()
                except AsignacionRevision.DoesNotExist:
                    pass 

            # ========= FIN DE LA TRANSACCIÓN ATÓMICA =========
            
            # Ahora, ejecutamos las funciones de recálculo:
            recalcular_costos_cargas_diesel() 
            
            if se_agrego_urea:
                recalcular_costos_cargas_urea()

            # Limpiar la sesión al finalizar exitosamente
            if diesel_data_key in self.request.session:
                del self.request.session[diesel_data_key]
                self.request.session.modified = True
            
            messages.success(self.request, f"¡Éxito! Proceso para {proceso.unidad.nombre} finalizado correctamente.")
            return redirect('encargado-pendientes-list')

        except Exception as e:
            # Capturar cualquier error inesperado durante el guardado y mostrarlo
            print(f"ERROR INESPERADO AL FINALIZAR EL PROCESO: {e}") # Log para el desarrollador
            messages.error(self.request, f"Error inesperado al guardar los datos: {e}. Por favor, contacte al administrador.")
            return self.form_invalid(form)
    
    
class UnidadDetailView(AdminRequiredMixin, DetailView):
    model = Unidad
    # Cambia el template por uno más específico para el reporte
    template_name = 'unidad_reporte_detalle.html' 

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        report_context = _get_unidad_report_context(self.request, self.object)
        context.update(report_context)
        unidad = self.get_object()

        # 1. Manejo del rango de fechas
        today = date.today()
        # Se obtiene el primer día del mes actual como fecha de inicio por defecto
        start_date_str = self.request.GET.get('start_date', (today.replace(day=1)).isoformat())
        end_date_str = self.request.GET.get('end_date', today.isoformat())
        
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)

        context['start_date'] = start_date
        context['end_date'] = end_date
        
        # 2. Filtrar todas las cargas por el rango de fechas
        context['cargas_diesel'] = CargaDiesel.objects.filter(
            unidad=unidad, fecha__date__range=[start_date, end_date]
        ).order_by('-fecha')
        
        context['cargas_urea'] = CargaUrea.objects.filter(
            unidad=unidad, fecha__date__range=[start_date, end_date]
        ).order_by('-fecha')

        # 3. Obtener el último registro de horas de thermo
        ultima_carga_con_thermo = CargaDiesel.objects.filter(
            unidad=unidad, hrs_thermo__isnull=False
        ).order_by('-fecha').first()
        context['ultimas_horas_thermo'] = ultima_carga_con_thermo.hrs_thermo if ultima_carga_con_thermo else "N/A"

        # 4. Procesar los checklists para encontrar fallas y fotos
        inspecciones = ChecklistInspeccion.objects.filter(
            unidad=unidad, fecha__date__range=[start_date, end_date]
        ).order_by('-fecha')
        
        fallas_checklist = []
        fotos_evidencia = []

        # Itera sobre cada campo del modelo de Checklist
        for inspeccion in inspecciones:
            for field in ChecklistInspeccion._meta.get_fields():
                # Revisa solo los campos que tienen opciones 'BIEN'/'MALO'
                if isinstance(field, models.CharField) and hasattr(field, 'choices') and field.choices:
                    valor = getattr(inspeccion, field.name)
                    if valor == 'MALO':
                        # Busca los campos de observación y foto correspondientes
                        obs_field_name = f"{field.name}_obs"
                        foto_field_name = f"{field.name}_foto"
                        
                        observacion = getattr(inspeccion, obs_field_name, "Sin observación.")
                        foto = getattr(inspeccion, foto_field_name, None)
                        
                        # Usa el verbose_name para una descripción amigable
                        nombre_amigable = field.verbose_name.title()

                        fallas_checklist.append({
                            'fecha': inspeccion.fecha,
                            'componente': nombre_amigable,
                            'observacion': observacion,
                        })

                        # === LÓGICA S3 ===
                        # El objeto 'foto' ahora es un FileField que podría
                        # tener una URL si está configurado S3
                        if foto and hasattr(foto, 'url') and foto.url:
                            fotos_evidencia.append({
                                'componente': nombre_amigable,
                                'url': foto.url,
                                'fecha': inspeccion.fecha
                            })
        
        context['fallas_checklist'] = fallas_checklist
        context['fotos_evidencia'] = fotos_evidencia
        context['titulo'] = f"Reporte Ejecutivo: {unidad.nombre}"
        
        return context
    
    
def _get_unidad_report_context(request, unidad):
    """
    Función auxiliar para obtener todo el contexto necesario para el reporte de una unidad.
    Puede ser utilizada tanto por la vista HTML como por la vista de generación de PDF.
    """
    context = {'unidad': unidad}
    
    # 1. Manejo del rango de fechas
    today = date.today()
    start_date_str = request.GET.get('start_date', (today.replace(day=1)).isoformat())
    end_date_str = request.GET.get('end_date', today.isoformat())
    
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)

    context['start_date'] = start_date
    context['end_date'] = end_date
    
    # 2. Filtrar todas las cargas por el rango de fechas
    context['cargas_diesel'] = CargaDiesel.objects.filter(
        unidad=unidad, fecha__date__range=[start_date, end_date]
    ).order_by('-fecha')
    
    context['cargas_urea'] = CargaUrea.objects.filter(
        unidad=unidad, fecha__date__range=[start_date, end_date]
    ).order_by('-fecha')

    # 3. Obtener el último registro de horas de thermo
    ultima_carga_con_thermo = CargaDiesel.objects.filter(
        unidad=unidad, hrs_thermo__isnull=False
    ).order_by('-fecha').first()
    context['ultimas_horas_thermo'] = ultima_carga_con_thermo.hrs_thermo if ultima_carga_con_thermo else "N/A"

    # 4. Procesar los checklists para encontrar fallas y fotos
    inspecciones = ChecklistInspeccion.objects.select_related('tecnico').filter(
        unidad=unidad, fecha__date__range=[start_date, end_date]
    ).order_by('-fecha')
    
    fallas_checklist = []
    fotos_evidencia = []

    for inspeccion in inspecciones:
        for field in ChecklistInspeccion._meta.get_fields():
            if isinstance(field, models.CharField) and hasattr(field, 'choices') and field.choices:
                valor = getattr(inspeccion, field.name)
                if valor == 'MALO':
                    obs_field_name = f"{field.name}_obs"
                    foto_field_name = f"{field.name}_foto"
                    
                    observacion = getattr(inspeccion, obs_field_name, "Sin observación.")
                    foto = getattr(inspeccion, foto_field_name, None)
                    
                    nombre_amigable = field.verbose_name.title()

                    fallas_checklist.append({
                        'fecha': inspeccion.fecha,
                        'componente': nombre_amigable,
                        'observacion': observacion,
                        'tecnico': inspeccion.tecnico.get_full_name() or inspeccion.tecnico.username,
                    })

                    # === LÓGICA S3 ===
                    if foto and hasattr(foto, 'url') and foto.url:
                        fotos_evidencia.append({
                            'componente': nombre_amigable,
                            'url': request.build_absolute_uri(foto.url),
                            'fecha': inspeccion.fecha
                        })
    
    context['fallas_checklist'] = fallas_checklist
    context['fotos_evidencia'] = fotos_evidencia
    context['titulo'] = f"Reporte Ejecutivo: {unidad.nombre}"
    
    # 5. Procesar inspecciones de llantas para encontrar alertas
    llantas_inspecciones = LlantasInspeccion.objects.select_related('tecnico').prefetch_related('detalles_llanta').filter(
        unidad=unidad, fecha__date__range=[start_date, end_date]
    ).order_by('-fecha')

    alertas_llantas = []
    for inspeccion in llantas_inspecciones:
        for detalle in inspeccion.detalles_llanta.all():
            status_mm = None
            status_presion = None

            # ========= INICIO: LÓGICA DE MM MODIFICADA =========
            if detalle.mm <= 5:
                status_mm = 'ROJO'
            elif 6 <= detalle.mm <= 7: # Incluye el 7
                status_mm = 'AMARILLO'
            # ========= FIN: LÓGICA DE MM MODIFICADA =========

            # Evaluar Presión
            if detalle.presion < 85:
                status_presion = 'ROJO'
            elif 85 <= detalle.presion <= 100:
                status_presion = 'AMARILLO'

            # Si hay alguna alerta para esta llanta, se agrega a la lista
            if status_mm or status_presion:
                alertas_llantas.append({
                    'fecha': inspeccion.fecha,
                    'posicion': detalle.posicion,
                    'mm': detalle.mm,
                    'presion': detalle.presion,
                    'status_mm': status_mm,
                    'status_presion': status_presion,
                    'tecnico': inspeccion.tecnico.get_full_name() or inspeccion.tecnico.username if inspeccion.tecnico else "N/A",
                })

    context['alertas_llantas'] = alertas_llantas
    
    return context

@login_required
def download_unidad_reporte_pdf(request, pk):
    """
    Vista que genera y sirve el reporte ejecutivo de una unidad como un archivo PDF.
    """
    if not es_admin(request.user):
        raise PermissionDenied("No tiene permiso para ver este reporte.")
    
    if HTML is None:
        return HttpResponse("Error: La librería WeasyPrint no está instalada. Contacte al administrador.", status=500)

    unidad = get_object_or_404(Unidad, pk=pk)
    context = _get_unidad_report_context(request, unidad)

    # Renderizar el template HTML a una cadena
    html_string = render_to_string('unidad_reporte_pdf.html', context)
    
    # Generar el PDF usando WeasyPrint
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    # Servir el PDF como una respuesta HTTP
    response = HttpResponse(pdf_file, content_type='application/pdf')
    filename = f"reporte_ejecutivo_{unidad.nombre.replace(' ', '_')}_{date.today()}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def download_checklist_excel(request, pk):
    """
    Genera y sirve un archivo Excel con el detalle de un Checklist de Inspección.
    """
    if not es_admin(request.user):
        raise PermissionDenied("No tiene permiso para exportar este reporte.")

    # 1. Obtener el objeto Checklist
    checklist = get_object_or_404(ChecklistInspeccion.objects.select_related(
        'unidad', 'operador', 'tecnico'
    ), pk=pk)

    # 2. Crear la respuesta HTTP con el tipo de contenido de Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # Define el nombre del archivo que se descargará
    filename = f"checklist_{checklist.unidad.nombre.replace(' ', '_')}_{checklist.fecha.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # 3. Crear el libro de trabajo y la hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle del Checklist"

    # 4. Definir estilos para los encabezados
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center')

    # 5. Agregar la información general del checklist
    ws['A1'] = 'Reporte de Checklist de Inspección'
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    ws.append([]) # Fila en blanco
    ws.append(['Unidad:', checklist.unidad.nombre])
    ws.append(['Operador:', str(checklist.operador)])
    ws.append(['Técnico:', checklist.tecnico.get_full_name() or checklist.tecnico.username])
    ws.append(['Fecha de Inspección:', checklist.fecha.strftime('%Y-%m-%d %H:%M:%S')])
    ws.append([]) # Fila en blanco

    # 6. Agregar la cabecera de la tabla de detalles
    header = ['Componente', 'Estado', 'Observaciones']
    ws.append(header)
    for cell in ws[ws.max_row]: # Aplica estilo a la última fila (la cabecera)
        cell.font = bold_font

    # 7. Iterar sobre los campos del modelo para poblar los datos
    # (Usando una lógica similar a la de la vista de creación/edición)
    field_groups = {
        'Estructura Exterior': ['cristales', 'espejos', 'logos', 'num_economico', 'puertas', 'cofre', 'parrilla', 'defensas', 'faros', 'plafoneria', 'stops', 'direccionales', 'tapiceria', 'instrumentos', 'carroceria', 'piso', 'costados', 'escape', 'pintura', 'franjas', 'loderas', 'extintor', 'senalamientos', 'estado_general'],
        'Mecánica y Motor': ['motor', 'caja', 'diferenciales', 'suspension_delantera', 'suspension_trasera', 'fugas_combustible', 'fugas_aceite', 'estado_llantas', 'presion_llantas', 'purga_tanques', 'estado_balatas', 'amortiguadores_delanteros', 'amortiguadores_traseros', 'rines_aluminio', 'mangueras_servicio', 'tarjeta_llave', 'revision_fusibles', 'revision_luces','revision_fuga_aire']
    }

    for group_name, field_list in field_groups.items():
        # Añadir una fila para el título del grupo
        ws.append([group_name])
        ws.merge_cells(f'A{ws.max_row}:C{ws.max_row}')
        ws.cell(row=ws.max_row, column=1).font = bold_font
        
        for field_name in field_list:
            field_obj = ChecklistInspeccion._meta.get_field(field_name)
            # Extraer el valor, la observación y el nombre amigable
            status = getattr(checklist, field_name, '')
            observation = getattr(checklist, f"{field_name}_obs", '')
            verbose_name = field_obj.verbose_name.title()
            
            # Añadir la fila con los datos
            ws.append([verbose_name, status, observation])

    # 8. Ajustar el ancho de las columnas para mejor legibilidad
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50

    # 9. Guardar el libro de trabajo en la respuesta HTTP
    wb.save(response)

    return response

class AdminPanelDeControlView(AdminOrEncargadoRequiredMixin, TemplateView):
    """
    Panel de control para administradores que muestra un resumen de todos
    los procesos que requieren atención en diferentes etapas.
    """
    template_name = 'admin_panel_de_control.html' # Usaremos un nuevo template

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 1. Procesos que completaron la primera parte (checklist y llantas) y 
        #    están esperando la carga de combustible por parte de un encargado/admin.
        context['procesos_pendientes_finalizar'] = ProcesoCarga.objects.filter(
            status='PENDIENTE'
        ).select_related('unidad', 'tecnico_inicia').order_by('fecha_inicio')

        # 2. Checklists que se crearon pero nunca se completó el segundo paso (llantas).
        #    Buscamos checklists que NO tienen un ProcesoCarga asociado.
        context['procesos_pendientes_llantas'] = ChecklistInspeccion.objects.filter(
            procesocarga__isnull=True
        ).select_related('unidad', 'tecnico').order_by('-fecha')
        
        context['titulo'] = "Panel de Control de Procesos"
        return context

@login_required
def admin_continuar_proceso_llantas(request, checklist_pk):
    """
    Permite a un administrador tomar control de un proceso que se quedó
    pendiente en la etapa de inspección de llantas.

    Esta vista establece el ID del checklist en la sesión del admin
    y lo redirige al formulario de llantas para completarlo.
    """
    if not es_admin(request.user):
        raise PermissionDenied("Solo los administradores pueden continuar procesos pendientes.")

    # Se busca el checklist que se quiere continuar.
    # Se asegura que el checklist realmente esté incompleto (no tiene un ProcesoCarga asociado).
    checklist = get_object_or_404(ChecklistInspeccion, pk=checklist_pk, procesocarga__isnull=True)

    # Se guarda el ID del checklist en la sesión del administrador.
    # Esto simula que el admin acaba de completar el paso anterior.
    request.session['proceso_checklist_id'] = checklist.id
    
    messages.info(request, f"Ha tomado el control del proceso para la unidad '{checklist.unidad.nombre}'. Por favor, complete la inspección de llantas.")
    
    # Redirige al administrador al formulario de llantas.
    return redirect('proceso-llantas', unidad_pk=checklist.unidad.pk)


class AjusteInventarioListView(AdminRequiredMixin, ListView):
    model = AjusteInventario
    template_name = 'generic_list.html'
    paginate_by = 25

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Ajustes de Inventario',
            'url_crear': 'ajusteinventario-create',
            'headers': ['#', 'Fecha', 'Suministro', 'Tipo', 'Cantidad (L)', 'Motivo', 'Responsable'],
        })
        return context

class AjusteInventarioCreateView(AdminRequiredMixin, CreateView):
    model = AjusteInventario
    form_class = AjusteInventarioForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('dashboard')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Ajuste de Inventario'
        return context

    def form_valid(self, form):
        form.instance.responsable = self.request.user
        messages.success(self.request, "Ajuste de inventario guardado correctamente.")
        return super().form_valid(form)
    

    
@login_required
def download_llantas_general_excel(request):
    """
    Genera y sirve un reporte en Excel con todas las inspecciones de llantas
    dentro de un rango de fechas, aplicando formato de tabla, centrado y condicional.
    """
    if not es_admin(request.user):
        raise PermissionDenied("No tiene permiso para exportar este reporte.")

    # 1. Obtener el rango de fechas de los parámetros GET
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    queryset = LlantasInspeccion.objects.select_related(
        'unidad', 'tecnico'
    ).prefetch_related('detalles_llanta').order_by('-fecha', 'unidad__nombre')

    if start_date_str and end_date_str:
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
        queryset = queryset.filter(fecha__date__range=[start_date, end_date])
        filename_date_part = f"{start_date.strftime('%Y%m%d')}_a_{end_date.strftime('%Y%m%d')}"
    else:
        filename_date_part = date.today().strftime('%Y%m%d')

    # 2. Crear la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"reporte_general_llantas_{filename_date_part}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # 3. Crear el libro y la hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte General de Llantas"

    # 4. Estilos
    bold_font = Font(bold=True)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 5. Encabezados de la tabla
    headers = [
        'Unidad', 'Fecha Inspección', 'Kilometraje', 'Técnico', 'Posición', 
        'MM', 'Marca', 'Modelo', 'Medida', 'Presión (PSI)'
    ]
    ws.append(headers)

    # 6. Poblar los datos
    for inspeccion in queryset:
        tecnico_nombre = inspeccion.tecnico.get_full_name() or inspeccion.tecnico.username if inspeccion.tecnico else "N/A"
        for detalle in inspeccion.detalles_llanta.all():
            ws.append([
                inspeccion.unidad.nombre,
                inspeccion.fecha.strftime('%Y-%m-%d %H:%M'),
                inspeccion.km,
                tecnico_nombre,
                detalle.posicion,
                detalle.mm,
                detalle.marca,
                detalle.modelo,
                detalle.medida,
                detalle.presion
            ])

    # 7. Obtener dimensiones y aplicar formato
    max_row = ws.max_row
    max_col_letra = 'J'
    
    if max_row > 1:
        # ========= INICIO: SECCIÓN CORREGIDA =========
        
        # Reglas para la columna de MM (F), aplicando solo a la columna F
        ws.conditional_formatting.add(f'F2:F{max_row}', CellIsRule(operator='lessThanOrEqual', formula=['5'], fill=red_fill))
        ws.conditional_formatting.add(f'F2:F{max_row}', CellIsRule(operator='between', formula=['6', '7'], fill=yellow_fill))
        ws.conditional_formatting.add(f'F2:F{max_row}', CellIsRule(operator='greaterThanOrEqual', formula=['8'], fill=green_fill))

        # Reglas para la columna de Presión (J), aplicando solo a la columna J
        ws.conditional_formatting.add(f'J2:J{max_row}', CellIsRule(operator='lessThan', formula=['85'], fill=red_fill))
        ws.conditional_formatting.add(f'J2:J{max_row}', CellIsRule(operator='between', formula=['85', '100'], fill=yellow_fill))
        ws.conditional_formatting.add(f'J2:J{max_row}', CellIsRule(operator='greaterThan', formula=['100'], fill=green_fill))
        
        # ========= FIN: SECCIÓN CORREGIDA =========
        
        # Crear la tabla
        table_range = f"A1:{max_col_letra}{max_row}"
        tab = Table(displayName="ReporteLlantas", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # Centrar todo el contenido de la tabla
        for row in ws[table_range]:
            for cell in row:
                cell.alignment = center_alignment

    # 8. Ajustar ancho de columnas
    column_widths = {'A': 15, 'B': 20, 'C': 15, 'D': 25, 'E': 15, 'F': 10, 'G': 20, 'H': 20, 'I': 15, 'J': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 9. Guardar y devolver
    wb.save(response)
    return response

class AsignacionRevisionView(AdminRequiredMixin, CreateView):
    model = AsignacionRevision
    form_class = AsignacionRevisionForm
    template_name = 'asignacion_revision_form.html'
    success_url = reverse_lazy('asignar-revision')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener la fecha para filtrar, o usar la fecha actual como default
        fecha_filtro_str = self.request.GET.get('fecha', timezone.now().strftime('%Y-%m-%d'))
        fecha_filtro = datetime.strptime(fecha_filtro_str, '%Y-%m-%d').date()

        # Obtener las asignaciones planeadas para ese día
        asignaciones_del_dia = AsignacionRevision.objects.filter(
            fecha_revision=fecha_filtro
        ).select_related('unidad')
        
        # --- INICIO DE LA LÓGICA CORREGIDA ---
        
        # Iterar sobre cada asignación para adjuntarle la info de fallas
        for asignacion in asignaciones_del_dia:
            
            # 1. Obtener TODAS las fallas PENDIENTES para la unidad de esta asignación
            fallas_pendientes = ChecklistCorreccion.objects.filter(
                inspeccion__unidad=asignacion.unidad,
                status='PENDIENTE'
            ).select_related('inspeccion') # Incluimos 'inspeccion' para acceder a ella

            # Crear una lista temporal para guardar los detalles de las fallas
            bad_items_list = []
            
            # 2. Iterar sobre las fallas pendientes encontradas
            for correccion_obj in fallas_pendientes:
                
                # --- Lógica para obtener la foto/obs de la inspección ORIGINAL ---
                campo_nombre = correccion_obj.nombre_campo
                # Obtenemos la inspección original (el ChecklistInspeccion)
                inspeccion_original = correccion_obj.inspeccion

                # Obtener la etiqueta (nombre amigable) desde la inspección original
                try:
                    label = inspeccion_original._meta.get_field(campo_nombre).verbose_name.title()
                except Exception:
                    label = campo_nombre.replace('_', ' ').title()

                # Definimos los nombres de los campos de obs y foto
                obs_field_name = f"{campo_nombre}_obs"
                foto_field_name = f"{campo_nombre}_foto"
                
                # Usamos getattr() para obtener la foto y la obs desde la INSPECCIÓN ORIGINAL
                observacion = getattr(inspeccion_original, obs_field_name, None)
                foto = getattr(inspeccion_original, foto_field_name, None)
                # --- Fin de la lógica de foto/obs ---

                # Añadimos la info a nuestra lista temporal
                bad_items_list.append({
                    'id': correccion_obj.id,
                    'label': label,
                    'obs': observacion or "Sin observación.",
                    'foto': foto, # <-- ¡Esta es la foto correcta!
                    'esta_corregido': correccion_obj.status == 'CORREGIDO', 
                    'comentario_admin': correccion_obj.comentario_admin or "",
                    'status': correccion_obj.get_status_display(),
                    'status_raw': correccion_obj.status,
                    'fecha_deteccion': inspeccion_original.fecha,
                })
            
            # --- ¡ESTA ES LA LÍNEA DE CORRECCIÓN IMPORTANTE! ---
            # Asignamos la lista que acabamos de construir de vuelta al
            # objeto 'asignacion' para que la plantilla HTML pueda leerla.
            asignacion.bad_items_list = bad_items_list
            # --- FIN DE LA CORRECCIÓN ---
            
            # Obtener la fecha del último checklist (para mostrar en el modal)
            latest_checklist = ChecklistInspeccion.objects.filter(
                unidad=asignacion.unidad
            ).order_by('-fecha').first()
            
            if latest_checklist:
                asignacion.latest_checklist_date = latest_checklist.fecha
            else:
                asignacion.latest_checklist_date = None
        
        # --- FIN DE LA LÓGICA CORREGIDA ---
        
        # Pasamos los datos al contexto del template
        context['asignaciones_del_dia'] = asignaciones_del_dia
        context['fecha_filtro'] = fecha_filtro
        context['titulo'] = "Asignar Revisiones de Unidades"
        
        # Añadir estadísticas (basado en el template 'asignacion_revision_form.html')
        stats = AsignacionRevision.objects.filter(
            fecha_revision=fecha_filtro
        ).aggregate(
            total=models.Count('id'),
            pendientes=models.Count('id', filter=models.Q(status='PENDIENTE')),
            en_proceso=models.Count('id', filter=models.Q(status='EN_PROCESO')),
            terminadas=models.Count('id', filter=models.Q(status='TERMINADO'))
        )
        context['stats'] = stats
        
        return context

    def form_valid(self, form):
        # Asignar el usuario que crea la asignación (basado en el modelo AsignacionRevision)
        # NOTA: Tu modelo 'AsignacionRevision' no tiene 'creado_por'.
        # Si lo añadieras, aquí iría: form.instance.creado_por = self.request.user
        messages.success(self.request, f"Unidad {form.instance.unidad} asignada correctamente para el {form.instance.fecha_revision.strftime('%d/%m/%Y')}.")
        return super().form_valid(form)

    def form_invalid(self, form):
        # Mejorar el mensaje de error para incluir el porqué (ej. duplicado)
        error_txt = form.errors.as_text()
        if 'unique_together' in error_txt:
             messages.error(self.request, f"No se pudo guardar: La unidad {form.cleaned_data.get('unidad')} ya tiene una revisión asignada para esa fecha.")
        else:
             messages.error(self.request, f"No se pudo guardar la asignación. Errores: {error_txt}")
        return redirect('asignar-revision')

@require_POST # Asegura que esta vista solo acepte peticiones POST
@login_required
def cancelar_revision(request, pk):
    if not es_admin(request.user):
        raise PermissionDenied

    asignacion = get_object_or_404(AsignacionRevision, pk=pk)
    
    nuevo_status = request.POST.get('status')
    comentario = request.POST.get('comentario', '')

    if nuevo_status in ['CANCELADO', 'NO_VINO']:
        asignacion.status = nuevo_status
        asignacion.comentario_cancelacion = comentario
        asignacion.save()
        messages.success(request, f"La revisión para la unidad {asignacion.unidad.nombre} ha sido marcada como '{asignacion.get_status_display()}'.")
    else:
        messages.error(request, "Estado de cancelación no válido.")

    # Redirige de vuelta al monitor del día de la asignación
    return redirect(f"{reverse('monitor-revisiones')}?fecha={asignacion.fecha_revision.strftime('%Y-%m-%d')}")

@login_required
@require_POST # Asegura que esta vista solo se pueda llamar con un método POST (desde un formulario)
def enviar_reporte_estado(request):
    """
    Vista que se encarga de llamar a la lógica para enviar
    el reporte de estado actual por correo y WhatsApp.
    """
    if not es_admin(request.user):
        messages.error(request, "No tiene permiso para realizar esta acción.")
        return redirect('dashboard')
        
    try:
        # Llamamos a la función que creamos y le pasamos el usuario que hizo la solicitud
        send_on_demand_status_report(request.user)
        messages.success(request, "¡Éxito! El reporte de estado actual ha sido enviado.")
    except Exception as e:
        messages.error(request, f"Ocurrió un error al intentar enviar el reporte: {e}")
        
    return redirect('dashboard')


@login_required
def download_cargadiesel_reporte(request):
    """
    Genera y sirve un reporte en Excel con el registro de todas las
    cargas de diésel (motor y thermo), incluyendo el tipo de unidad.
    """
    if not es_admin(request.user):
        raise PermissionDenied("No tiene permiso para exportar este reporte.")

    # 1. Obtener los datos (aplicando los mismos filtros de la vista CargaDieselListView)
    # Usamos select_related('unidad') para optimizar la consulta
    queryset = CargaDiesel.objects.select_related('unidad').order_by('-fecha')
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    unidad_id = request.GET.get('unidad')

    if start_date_str and end_date_str:
        # Asegurarse de que las fechas no estén vacías antes de filtrar
        if start_date_str and end_date_str:
             queryset = queryset.filter(fecha__date__range=[date.fromisoformat(start_date_str), date.fromisoformat(end_date_str)])
    
    if unidad_id:
        queryset = queryset.filter(unidad_id=unidad_id)

    # 2. Crear la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"reporte_cargas_diesel_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # 3. Crear el libro y la hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Cargas Diesel"

    # 4. Estilos
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 5. Encabezados de la tabla
    headers = [
        'Fecha de Carga', 'Unidad', 'Tipo de Unidad', 
        'Lts Motor', 'Lts Thermo', 'Total Litros'
    ]
    ws.append(headers)
    
    # Aplicar estilo a la cabecera
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = center_alignment

    # 6. Poblar los datos
    for carga in queryset:
        # Formatear la fecha
        fecha_carga = carga.fecha.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M')
        
        # Obtener datos de la carga
        lts_motor = carga.lts_diesel or 0
        lts_thermo = carga.lts_thermo or 0
        total_litros = lts_motor + lts_thermo
        
        # Obtener datos de la unidad (ya precargada con select_related)
        unidad_nombre = carga.unidad.nombre
        tipo_legible = carga.unidad.get_tipo_display()
        
        ws.append([
            fecha_carga,
            unidad_nombre,
            tipo_legible,
            lts_motor,
            lts_thermo,
            total_litros
        ])

    # 7. Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25  # Fecha
    ws.column_dimensions['B'].width = 25  # Unidad
    ws.column_dimensions['C'].width = 20  # Tipo de Unidad
    ws.column_dimensions['D'].width = 15  # Lts Motor
    ws.column_dimensions['E'].width = 15  # Lts Thermo
    ws.column_dimensions['F'].width = 15  # Total Litros

    # 8. Guardar y devolver
    wb.save(response)
    return response

@login_required
def download_unidades_excel(request):
    """
    Genera y sirve un reporte en Excel con el listado de unidades,
    su tipo, KM actual y los detalles de su última carga de diésel.
    """
    if not es_admin(request.user):
        raise PermissionDenied("No tiene permiso para exportar este reporte.")

    # 1. Obtener los datos (aplicando los mismos filtros de la vista de lista)
    queryset = Unidad.objects.order_by('nombre')
    search_query = request.GET.get('q')
    tipo_filter = request.GET.get('tipo_unidad')

    if search_query:
        queryset = queryset.filter(nombre__icontains=search_query)
    if tipo_filter:
        queryset = queryset.filter(tipo=tipo_filter)

    # 2. Crear la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"reporte_unidades_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # 3. Crear el libro y la hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Unidades"

    # 4. Estilos
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 5. Encabezados de la tabla
    headers = [
        'Unidad', 'Tipo de Unidad', 'KM Actual', 
        'Fecha Última Carga', 'Lts Motor (Últ Carga)', 
        'Lts Thermo (Últ Carga)', 'Total Litros (Últ Carga)'
    ]
    ws.append(headers)
    
    # Aplicar estilo a la cabecera
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = center_alignment

    # 6. Poblar los datos
    for unidad in queryset:
        tipo_legible = unidad.get_tipo_display()
        
        # Buscar la última carga de diésel para esta unidad
        ultima_carga = CargaDiesel.objects.filter(unidad=unidad).order_by('-fecha').first()
        
        # Inicializar valores por defecto
        fecha_ultima_carga = "N/A"
        lts_motor = 0
        lts_thermo = 0
        total_litros = 0

        if ultima_carga:
            # Si se encontró una carga, obtener sus datos
            fecha_ultima_carga = ultima_carga.fecha.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M')
            lts_motor = ultima_carga.lts_diesel or 0
            lts_thermo = ultima_carga.lts_thermo or 0
            total_litros = lts_motor + lts_thermo
        
        ws.append([
            unidad.nombre,
            tipo_legible,
            unidad.km_actual,
            fecha_ultima_carga,
            lts_motor,
            lts_thermo,
            total_litros
        ])

    # 7. Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25  # Unidad
    ws.column_dimensions['B'].width = 20  # Tipo de Unidad
    ws.column_dimensions['C'].width = 15  # KM Actual
    ws.column_dimensions['D'].width = 25  # Fecha Última Carga
    ws.column_dimensions['E'].width = 20  # Lts Motor
    ws.column_dimensions['F'].width = 20  # Lts Thermo
    ws.column_dimensions['G'].width = 20  # Total Litros

    # 8. Guardar y devolver
    wb.save(response)
    return response


@login_required
def corregir_checklist_mal_view(request):
    
    # 1. Definir el FormSet basado en el formulario de corrección
    ChecklistCorreccionFormSet = modelformset_factory(
        ChecklistCorreccion,
        form=ChecklistCorreccionForm,
        extra=0, # No se pueden añadir nuevos registros
        can_delete=False
    )
    
    # 2. Obtener los ítems MALO pendientes de corregir
    queryset = ChecklistCorreccion.objects.filter(
        status='PENDIENTE'
    ).select_related(
        'inspeccion', 
        'inspeccion__unidad',
        'inspeccion__tecnico' # <-- Añadir 'inspeccion__tecnico'
    ).order_by(
        'inspeccion__fecha', # Segregado por fecha de inspección
        'inspeccion__unidad__nombre'
    )
    
    # --- INICIO: LÓGICA DE FILTROS GET MODIFICADA ---
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    unidad_id = request.GET.get('unidad')
    tecnico_id = request.GET.get('tecnico') # <-- NUEVA LÍNEA
    
    selected_unidad = None
    selected_tecnico = None # <-- NUEVA LÍNEA

    if start_date_str and end_date_str:
        try:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
            queryset = queryset.filter(inspeccion__fecha__date__range=[start_date, end_date])
        except ValueError:
            start_date_str = '' 
            end_date_str = ''
    
    if unidad_id:
        queryset = queryset.filter(inspeccion__unidad_id=unidad_id)
        selected_unidad = Unidad.objects.filter(pk=unidad_id).first()

    # --- NUEVO: Procesar filtro de técnico ---
    if tecnico_id:
        queryset = queryset.filter(inspeccion__tecnico_id=tecnico_id)
        selected_tecnico = User.objects.filter(pk=tecnico_id).first()
    # --- FIN: LÓGICA DE FILTROS GET MODIFICADA ---
    
    
    if request.method == 'POST':
        formset = ChecklistCorreccionFormSet(request.POST, queryset=queryset)
        
        if formset.is_valid():
            with transaction.atomic():
                correcciones_count = 0
                descartes_count = 0
                comentarios_count = 0
                
                for form in formset:
                    if form.has_changed():
                        instance = form.save(commit=False)
                        
                        marcar_corregido = form.cleaned_data.get('marcar_corregido')
                        marcar_descartado = form.cleaned_data.get('marcar_descartado')
                        inspeccion_original = instance.inspeccion
                        
                        if marcar_corregido:
                            instance.status = 'CORREGIDO'
                            instance.comentario_admin = form.cleaned_data.get('comentario_admin', '')
                            instance.corregido_por = request.user 
                            instance.fecha_correccion = timezone.now()
                            instance.save() 
                            
                            setattr(inspeccion_original, instance.nombre_campo, 'BIEN')
                            inspeccion_original.save() 
                            correcciones_count += 1
                        
                        elif marcar_descartado:
                            instance.status = 'DESCARTADO'
                            instance.comentario_admin = form.cleaned_data.get('comentario_admin', '')
                            instance.corregido_por = request.user 
                            instance.fecha_correccion = timezone.now() 
                            instance.save() 
                            
                            setattr(inspeccion_original, instance.nombre_campo, 'BIEN')
                            inspeccion_original.save() 
                            descartes_count += 1

                        elif 'comentario_admin' in form.changed_data:
                            instance.save()
                            comentarios_count += 1

                messages.success(request, f"Proceso completado: {correcciones_count} ítems corregidos, {descartes_count} ítems descartados, {comentarios_count} comentarios actualizados.")
                
                return redirect(request.get_full_path())
        else:
            messages.error(request, "Error en el formulario. Por favor, revisa los datos.")
    else:
        formset = ChecklistCorreccionFormSet(queryset=queryset)

    # 3. Lógica de Agrupación por Fecha para el Template
    items_agrupados = {}
    form_map = {form.instance.id: form for form in formset} 
    
    for detalle in queryset:
        fecha_inspeccion = detalle.inspeccion.fecha.date() 
        if fecha_inspeccion not in items_agrupados:
            items_agrupados[fecha_inspeccion] = []
        
        detalle.form = form_map.get(detalle.id)
        detalle.etiqueta_legible = detalle.inspeccion._meta.get_field(detalle.nombre_campo).verbose_name or detalle.nombre_campo.replace('_', ' ').title()
        
        foto_obj = getattr(detalle.inspeccion, f"{detalle.nombre_campo}_foto", None)
        detalle.foto = foto_obj
        
        items_agrupados[fecha_inspeccion].append(detalle)

    data_for_template = [
        {'fecha': fecha, 'detalles': detalles} 
        for fecha, detalles in items_agrupados.items()
    ]
    
    context = {
        'titulo': 'Panel de Corrección de Checklist (MALO)',
        'items_agrupados': data_for_template,
        'formset': formset,
        # --- Pasar nuevos filtros a la plantilla ---
        'start_date': start_date_str,
        'end_date': end_date_str,
        'selected_unidad': selected_unidad,
        'selected_tecnico': selected_tecnico, # <-- NUEVA LÍNEA
    }
    
    return render(request, 'flota/corregir_checklist_mal.html', context)

class MonitorRevisionesView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    # La lógica de carga de datos se maneja en get_context_data
    template_name = 'monitor_revisiones.html'
    context_object_name = 'revision_items'
    
    # 1. Función de prueba para asegurar que solo usuarios con permiso puedan acceder
    def test_func(self):
        # Asumo que el grupo de 'Administradores' o ser 'staff' es el requisito
        return self.request.user.groups.filter(name='Administradores').exists() or self.request.user.is_staff

    # 2. Obtener y mapear los datos
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener la fecha de filtro de la URL o usar la fecha local actual
        fecha_str = self.request.GET.get('fecha')
        fecha_filtro = parse_date(fecha_str) if fecha_str else timezone.localdate()

        context['fecha_filtro'] = fecha_filtro
        context['titulo'] = f'Monitor de Revisiones Programadas - {fecha_filtro.strftime("%d/%m/%Y")}'
        
        # 1. Obtener las ASIGNACIONES (el plan) para esa fecha
        asignaciones = AsignacionRevision.objects.filter(
            fecha_revision=fecha_filtro
        ).select_related(
            'unidad'
        ).order_by(
            'unidad__nombre'
        )

        # 2. Obtener los PROCESOS (la ejecución) de ese día
        # (Este es el paso que faltaba y estaba incorrecto)
        procesos_del_dia = ProcesoCarga.objects.filter(
            fecha_inicio__date=fecha_filtro
        ).select_related(
            'tecnico_inicia',
            'checklist',
            'inspeccion_llantas',
            'carga_diesel',
            'carga_urea'
        )
        
        # 3. Mapear los procesos a sus unidades para un acceso rápido
        mapa_procesos = {p.unidad_id: p for p in procesos_del_dia}

        # 4. Combinar la data para el template
        revision_items = []
        for asignacion in asignaciones:
            # Usamos el mapa para encontrar el proceso ejecutado (si existe)
            proceso = mapa_procesos.get(asignacion.unidad_id)
            
            tiempo_total = None
            if proceso and proceso.status == 'COMPLETADO' and proceso.fecha_fin:
                delta = proceso.fecha_fin - proceso.fecha_inicio
                # Formatea el tiempo (ej: "0:35:10")
                tiempo_total = str(timedelta(seconds=int(delta.total_seconds())))

            # Pasa los datos de progreso al template
            revision_items.append({
                'asignacion': asignacion, # El objeto AsignacionRevision
                'proceso': proceso,       # El objeto ProcesoCarga (o None)
                'tiempo_total': tiempo_total,
                # Banderas para los íconos
                'checklist_ok': bool(proceso and proceso.checklist),
                'llantas_ok': bool(proceso and proceso.inspeccion_llantas),
                'diesel_ok': bool(proceso and proceso.carga_diesel),
                'urea_ok': bool(proceso and (proceso.carga_urea or proceso.status == 'COMPLETADO')),
            })
            
        # El template itera sobre 'monitor_data'
        context['monitor_data'] = revision_items
        return context

    # 4. Método requerido por ListView
    def get_queryset(self):
        # Este método solo sirve como base para ListView. La lógica clave está en get_context_data.
        fecha_str = self.request.GET.get('fecha')
        fecha_filtro = parse_date(fecha_str) if fecha_str else timezone.localdate()
        return AsignacionRevision.objects.filter(fecha_revision=fecha_filtro).order_by('unidad__nombre')
    
@login_required
def download_comprasuministro_excel(request):
    """
    Genera y devuelve un archivo Excel con el listado de Compras de Suministros, 
    aplicando los filtros de fecha y tipo_suministro.
    """
    # 1. Obtener los filtros del GET request
    tipo_filter = request.GET.get('tipo_suministro')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # 2. Construir el QuerySet
    queryset = CompraSuministro.objects.all().order_by('-fecha_compra')
    
    if tipo_filter:
        queryset = queryset.filter(tipo_suministro=tipo_filter)

    if start_date_str and end_date_str:
        try:
            # Asegúrese de tener 'date' importado de 'datetime'
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
            queryset = queryset.filter(fecha_compra__date__range=[start_date, end_date])
        except ValueError:
            # Manejar error si el formato de fecha es incorrecto
            pass

    # 3. Crear el libro y la hoja de trabajo de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Listado_Compras_Suministros.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras_Suministros"
    
    # 4. Cabeceras
    headers = [
        "ID", 
        "Fecha de Compra", 
        "Tipo de Suministro", 
        "Cantidad Comprada", 
        "Costo Total ($)", 
        "Precio por Litro/Unidad ($)",
        "Proveedor", 
        "Litros/Unidades Restantes"
    ]
    ws.append(headers)

    # Aplicar formato a la cabecera
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 5. Escribir los datos
    row_num = 2
    for compra in queryset:
        # Formatear la fecha para mejor visualización
        fecha_str = timezone.localtime(compra.fecha_compra).strftime('%Y-%m-%d %H:%M') if compra.fecha_compra else ''
        
        row = [
            compra.pk,
            fecha_str,
            compra.get_tipo_suministro_display(),
            compra.cantidad,
            # ========= INICIO DE LA CORRECCIÓN =========
            compra.precio, # <--- Cambia 'compra.costo' por 'compra.precio'
            # ========= FIN DE LA CORRECCIÓN =========
            compra.precio_por_litro,
            compra.proveedor,
            compra.litros_restantes,
        ]
        ws.append(row)
        # Aplicar formato a los valores numéricos
        ws[f'D{row_num}'].number_format = '#,##0.00'  # Cantidad
        ws[f'E{row_num}'].number_format = '$#,##0.00' # Costo Total
        ws[f'F{row_num}'].number_format = '$#,##0.00' # Precio/Unidad
        ws[f'H{row_num}'].number_format = '#,##0.00'  # Restantes
        
        row_num += 1

    # 6. Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 7. Guardar el archivo y devolver la respuesta
    wb.save(response)
    return response

@login_required
@require_POST # Esta vista solo acepta peticiones POST
def api_corregir_falla_checklist(request, pk):
    """
    API interna para guardar una corrección de checklist desde el modal
    usando AJAX.
    (MODIFICADA para aceptar 'marcar_corregido' y 'marcar_descartado')
    """
    if not (es_admin(request.user) or es_encargado(request.user)):
        return JsonResponse({'status': 'error', 'message': 'No tienes permiso.'}, status=403)

    try:
        # 1. Encontrar el objeto de corrección
        correccion = get_object_or_404(ChecklistCorreccion, pk=pk)
        
        # 2. Leer los datos enviados
        data = json.loads(request.body)
        marcar_corregido = data.get('marcar_corregido', False)
        marcar_descartado = data.get('marcar_descartado', False)
        comentario_admin = data.get('comentario_admin', '')

        # 3. Validar (que no sean ambos)
        if marcar_corregido and marcar_descartado:
            return JsonResponse({
                'status': 'error',
                'message': 'No puede marcar un ítem como "Corregido" y "Descartado" al mismo tiempo.'
            }, status=400)

        # 4. Aplicar la lógica de guardado
        with transaction.atomic():
            if marcar_corregido and correccion.status == 'PENDIENTE':
                # LÓGICA 1: Se marcó como CORREGIDO
                correccion.status = 'CORREGIDO'
                correccion.comentario_admin = comentario_admin
                correccion.corregido_por = request.user
                correccion.fecha_correccion = timezone.now()
                
                inspeccion_original = correccion.inspeccion
                setattr(inspeccion_original, correccion.nombre_campo, 'BIEN')
                inspeccion_original.save() 
                correccion.save()

            elif marcar_descartado and correccion.status == 'PENDIENTE':
                # LÓGICA 2: Se marcó como DESCARTADO
                correccion.status = 'DESCARTADO'
                correccion.comentario_admin = comentario_admin
                correccion.corregido_por = request.user 
                correccion.fecha_correccion = timezone.now()
                
                inspeccion_original = correccion.inspeccion
                setattr(inspeccion_original, correccion.nombre_campo, 'BIEN')
                inspeccion_original.save()
                correccion.save()

            elif not marcar_corregido and not marcar_descartado:
                # LÓGICA 3: Solo se cambió el comentario
                correccion.comentario_admin = comentario_admin
                correccion.save()
            
            # (Si ya estaba corregido/descartado, no hacer nada más que guardar comentario)
            elif correccion.status != 'PENDIENTE':
                 correccion.comentario_admin = comentario_admin
                 correccion.save()


        return JsonResponse({
            'status': 'ok',
            'message': 'Corrección guardada exitosamente.',
            'fecha_correccion': correccion.fecha_correccion.strftime('%d/%m/%Y') if correccion.fecha_correccion else None,
            'corregido_por': correccion.corregido_por.get_full_name() if correccion.corregido_por else None,
            'new_status': correccion.get_status_display(),
        })

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Datos inválidos.'}, status=400)
    except ChecklistCorreccion.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'El ítem no existe.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
@login_required
def download_correcciones_report_pdf(request, pk):
    if not es_admin(request.user):
        raise PermissionDenied("No tiene permiso para ver este reporte.")
    if HTML is None:
        return HttpResponse("Error: La librería WeasyPrint no está instalada.", status=500)

    unidad = get_object_or_404(Unidad, pk=pk)
    today = timezone.localdate()
    start_date_str = request.GET.get('start_date', (today.replace(day=1)).isoformat())
    end_date_str = request.GET.get('end_date', today.isoformat())
    
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)

    queryset = ChecklistCorreccion.objects.filter(
        inspeccion__unidad=unidad,
        esta_corregido=True,
        fecha_correccion__date__range=[start_date, end_date]
    ).select_related(
        'inspeccion', 'corregido_por',
        'inspeccion__tecnico' # <-- AÑADIDO: Para obtener el técnico que reportó
    ).order_by('-fecha_correccion')

    correcciones_procesadas = []
    for item in queryset:
        fecha_falla = item.inspeccion.fecha
        fecha_correccion = item.fecha_correccion
        dias_totales = "N/A"
        if fecha_falla and fecha_correccion:
            dias_totales = (fecha_correccion.date() - fecha_falla.date()).days
        
        try:
            componente = item.inspeccion._meta.get_field(item.nombre_campo).verbose_name.title()
        except Exception:
            componente = item.nombre_campo.replace('_', ' ').title()
        
        foto_obj = getattr(item.inspeccion, f"{item.nombre_campo}_foto", None)
        
        # Obtenemos el técnico que reportó
        tecnico_reporta = "N/A"
        if item.inspeccion.tecnico:
            tecnico_reporta = item.inspeccion.tecnico.get_full_name() or item.inspeccion.tecnico.username

        correcciones_procesadas.append({
            'componente': componente,
            'tecnico_reporta': tecnico_reporta, # <-- AÑADIDO
            'fecha_falla': fecha_falla,
            'fecha_correccion': fecha_correccion,
            'dias_totales': dias_totales,
            'observacion_original': item.observacion_original,
            'comentario_admin': item.comentario_admin,
            'corregido_por': item.corregido_por.get_full_name() if item.corregido_por else "N/A",
            'foto_url': request.build_absolute_uri(foto_obj.url) if foto_obj and foto_obj.url else None,
        })

    context = {
        'unidad': unidad,
        'start_date': start_date,
        'end_date': end_date,
        'correcciones_list': correcciones_procesadas,
        'titulo': f"Reporte de Fallas Corregidas: {unidad.nombre}",
        'today': today, # <-- AÑADIDO: Para el header del PDF
    }

    html_string = render_to_string('reporte_correcciones_pdf.html', context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    filename = f"reporte_correcciones_{unidad.nombre.replace(' ', '_')}_{today.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# =====================================================================
# === VISTA DEL EXCEL DE CORRECCIONES (NUEVA) =========================
# =====================================================================
@login_required
def download_correcciones_report_excel(request, pk):
    if not es_admin(request.user):
        raise PermissionDenied("No tiene permiso para exportar este reporte.")

    unidad = get_object_or_404(Unidad, pk=pk)
    today = timezone.localdate()
    start_date_str = request.GET.get('start_date', (today.replace(day=1)).isoformat())
    end_date_str = request.GET.get('end_date', today.isoformat())
    
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)

    queryset = ChecklistCorreccion.objects.filter(
        inspeccion__unidad=unidad,
        status='CORREGIDO', # <-- CAMBIO: de esta_corregido=True a status='CORREGIDO'
        fecha_correccion__date__range=[start_date, end_date]
    ).select_related(
        'inspeccion', 'corregido_por', 'inspeccion__tecnico'
    ).order_by('-fecha_correccion')

    # --- Inicio de la creación del Excel ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"reporte_correcciones_{unidad.nombre.replace(' ', '_')}_{today.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Correcciones"

    # Estilos
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    # Título
    ws['A1'] = f"Reporte de Fallas Corregidas: {unidad.nombre}"
    ws.merge_cells('A1:I1')
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align

    ws['A2'] = f"Reporte generado el {today.strftime('%d/%m/%Y')}. Rango: {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}"
    ws.merge_cells('A2:I2')
    ws['A2'].alignment = center_align
    ws.append([]) # Fila vacía

    # Encabezados
    headers = [
        'Componente', 'Técnico (Reportó)', 'Fecha Falla', 'Comentario Falla',
        'Fecha Corrección', 'Días en Reparar', 'Comentario Admin', 
        'Corregido Por', 'URL Foto Evidencia'
    ]
    ws.append(headers)
    for cell in ws[4]: # Fila 4 es la de headers
        cell.font = bold_font

    # Llenar datos
    for item in queryset:
        fecha_falla = item.inspeccion.fecha
        fecha_correccion = item.fecha_correccion
        dias_totales = "N/A"
        if fecha_falla and fecha_correccion:
            dias_totales = (fecha_correccion.date() - fecha_falla.date()).days
        
        try:
            componente = item.inspeccion._meta.get_field(item.nombre_campo).verbose_name.title()
        except Exception:
            componente = item.nombre_campo.replace('_', ' ').title()
        
        foto_obj = getattr(item.inspeccion, f"{item.nombre_campo}_foto", None)
        foto_url = request.build_absolute_uri(foto_obj.url) if foto_obj and foto_obj.url else "Sin Foto"

        tecnico_reporta = "N/A"
        if item.inspeccion.tecnico:
            tecnico_reporta = item.inspeccion.tecnico.get_full_name() or item.inspeccion.tecnico.username

        corregido_por = item.corregido_por.get_full_name() if item.corregido_por else "N/A"

        # CORRECCIÓN 2: Quitar la información de zona horaria (tzinfo)
        fecha_falla_naive = timezone.localtime(fecha_falla).replace(tzinfo=None) if fecha_falla else "N/A"
        fecha_correccion_naive = timezone.localtime(fecha_correccion).replace(tzinfo=None) if fecha_correccion else "N/A"

        # Añadir fila
        ws.append([
            componente,
            tecnico_reporta,
            fecha_falla_naive,
            item.observacion_original or "-",
            fecha_correccion_naive,
            dias_totales,
            item.comentario_admin or "-",
            corregido_por,
            foto_url
        ])
        
        # Formatear fechas
        last_row = ws.max_row
        if fecha_falla:
            ws[f'C{last_row}'].number_format = 'DD/MM/YYYY HH:MM'
        if fecha_correccion:
            ws[f'E{last_row}'].number_format = 'DD/MM/YYYY HH:MM'

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30 # Componente
    ws.column_dimensions['B'].width = 30 # Técnico
    ws.column_dimensions['C'].width = 20 # Fecha Falla
    ws.column_dimensions['D'].width = 40 # Comentario Falla
    ws.column_dimensions['E'].width = 20 # Fecha Corrección
    ws.column_dimensions['F'].width = 15 # Días
    ws.column_dimensions['G'].width = 40 # Comentario Admin
    ws.column_dimensions['H'].width = 30 # Corregido Por
    ws.column_dimensions['I'].width = 50 # URL Foto

    # Guardar
    wb.save(response)
    return response

class EntregaSuministrosListView(AdminRequiredMixin, ListView):
    model = EntregaSuministros
    template_name = 'flota/entrega_suministros_list.html' 
    context_object_name = 'object_list'
    paginate_by = 25

    def get_queryset(self):
        queryset = super().get_queryset().select_related('operador', 'unidad').order_by('-fecha_entrega')
        
        operador_id = self.request.GET.get('operador')
        unidad_id = self.request.GET.get('unidad') 
        
        if operador_id:
            queryset = queryset.filter(operador_id=operador_id)
        if unidad_id:
            queryset = queryset.filter(unidad_id=unidad_id) 
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # --- ENCABEZADOS ACTUALIZADOS ---
        context.update({
            'titulo': 'Entregas de Cinchos y Tarjetas',
            'url_crear': 'entrega-suministros-create',
            'headers': ['#', 'Operador', 'Fecha', 'Cant. Cinchos', 'F. Inicial', 'F. Final', 'Unidad', 'Motor', 'Thermo', 'Entregado'],
            'url_update_name': 'entrega-suministros-update',
            'url_delete_name': 'entrega-suministros-delete',
        })
        
        if selected_operador_id := self.request.GET.get('operador'):
            context['selected_operador'] = Operador.objects.filter(pk=selected_operador_id).first()
        if selected_unidad_id := self.request.GET.get('unidad'):
            context['selected_unidad'] = Unidad.objects.filter(pk=selected_unidad_id).first()
            
        return context
class EntregaSuministrosCreateView(AdminRequiredMixin, CreateView):
    model = EntregaSuministros
    form_class = EntregaSuministrosForm
    
    # --- CORRIGE ESTA LÍNEA ---
    template_name = 'flota/entrega_suministros_form.html' # Antes decía 'generic_form.html'
    # --- FIN DE LA CORRECCIÓN ---

    success_url = reverse_lazy('entrega-suministros-list')

    def form_valid(self, form):
        form.instance.entregado_por = self.request.user
        messages.success(self.request, "Entrega registrada correctamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Registrar Nueva Entrega de Suministros'
        return context

class EntregaSuministrosUpdateView(AdminRequiredMixin, UpdateView):
    model = EntregaSuministros
    form_class = EntregaSuministrosForm
    
    # --- CORRIGE ESTA LÍNEA ---
    template_name = 'flota/entrega_suministros_form.html' # Antes decía 'generic_form.html'
    # --- FIN DE LA CORRECCIÓN ---

    success_url = reverse_lazy('entrega-suministros-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f"Editar Entrega a {self.object.operador}"
        return context

class EntregaSuministrosDeleteView(AdminRequiredMixin, DeleteView):
    model = EntregaSuministros
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('entrega-suministros-list')
    
    def form_valid(self, form):
        messages.success(self.request, f"Entrega a '{self.object.operador}' eliminada exitosamente.")
        return super().form_valid(form)
    
@login_required
def search_operadores_api(request):
    term = request.GET.get('term', '')
    operadores = Operador.objects.filter(Q(nombre__icontains=term) | Q(apellido__icontains=term))[:15]
    results = [{'id': o.id, 'text': f"{o.nombre} {o.apellido}"} for o in operadores]
    return JsonResponse({'results': results}) #121