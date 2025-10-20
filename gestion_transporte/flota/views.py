# flota/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.core.exceptions import PermissionDenied
from datetime import date, timedelta
from decimal import Decimal
from django.views.generic.edit import FormView
from django.db import transaction
from django.contrib import messages
from django.forms import formset_factory, inlineformset_factory
from .models import *
from .forms import *
import openpyxl
from openpyxl.styles import Font, Alignment

# -----------------------------------------------------------------------------
# 1. MIXINS Y VISTAS DE LÓGICA
# -----------------------------------------------------------------------------

def es_admin(user):
    return user.is_authenticated and user.groups.filter(name='Administrador').exists()

def es_tecnico(user):
    return user.is_authenticated and user.groups.filter(name='Tecnico').exists()

# --- CAMBIO 1: Nueva vista "Dispatcher" para la página de inicio ---
# Esta será la nueva página de inicio. Es inteligente y sabe qué hacer con cada rol.
@login_required
def home_dispatcher_view(request):
    if es_admin(request.user):
        # Si es admin, lo enviamos a la vista del dashboard de admin.
        return AdminDashboardView.as_view()(request)
    elif es_tecnico(request.user):
        # Si es técnico, lo redirigimos a su flujo de trabajo.
        return redirect('tecnico-seleccionar-unidad')
    else:
        # Si no tiene un rol, negamos el acceso.
        raise PermissionDenied("No tienes un rol asignado para acceder a esta aplicación.")

class AdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return es_admin(self.request.user)

class TecnicoRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return es_tecnico(self.request.user)


# -----------------------------------------------------------------------------
# 2. VISTA PRINCIPAL (DASHBOARD)
# -----------------------------------------------------------------------------

class AdminDashboardView(AdminRequiredMixin, ListView):
    model = Unidad
    template_name = 'dashboard.html'
    context_object_name = 'unidades'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # --- LÓGICA DE INVENTARIO TOTAL (SIN CAMBIOS) ---
        total_diesel_comprado = CompraSuministro.objects.filter(tipo_suministro='DIESEL').aggregate(total=Sum('cantidad'))['total'] or 0
        total_urea_comprada = CompraSuministro.objects.filter(tipo_suministro='UREA').aggregate(total=Sum('cantidad'))['total'] or 0
        total_aceite_comprado = CompraSuministro.objects.filter(tipo_suministro='ACEITE').aggregate(total=Sum('cantidad'))['total'] or 0
        consumo_diesel_total = CargaDiesel.objects.aggregate(total=Sum('lts_diesel'))['total'] or 0
        consumo_thermo_total = CargaDiesel.objects.aggregate(total=Sum('lts_thermo'))['total'] or 0
        consumo_urea_total = CargaUrea.objects.aggregate(total=Sum('litros_cargados'))['total'] or 0
        consumo_aceite_total = CargaAceite.objects.aggregate(total=Sum('cantidad'))['total'] or 0
        
        context['diesel_restante'] = total_diesel_comprado - (consumo_diesel_total + consumo_thermo_total)
        context['urea_restante'] = total_urea_comprada - consumo_urea_total
        context['aceite_restante'] = total_aceite_comprado - consumo_aceite_total

        # --- LÓGICA PARA FILTRO DE FECHAS ---
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        if start_date_str and end_date_str:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
        else:
            today = date.today()
            last_sunday = today - timedelta(days=(today.weekday() + 1))
            last_monday = last_sunday - timedelta(days=6)
            start_date = last_monday
            end_date = last_sunday
            
        context['start_date'] = start_date
        context['end_date'] = end_date

        # --- NUEVA LÓGICA PARA OBTENER CONSUMO POR UNIDAD EN RANGO ---
        unidades = Unidad.objects.all()
        
        # 1. Obtener datos de consumo agrupados por unidad de forma eficiente
        consumo_diesel = CargaDiesel.objects.filter(fecha__date__range=[start_date, end_date]) \
            .values('unidad_id') \
            .annotate(total_motor=Sum('lts_diesel'), total_thermo=Sum('lts_thermo'))
            
        consumo_urea = CargaUrea.objects.filter(fecha__date__range=[start_date, end_date]) \
            .values('unidad_id') \
            .annotate(total_urea=Sum('litros_cargados'))

        consumo_aceite = CargaAceite.objects.filter(fecha__date__range=[start_date, end_date]) \
            .values('unidad_id') \
            .annotate(total_aceite=Sum('cantidad'))

        # 2. Convertir los datos a diccionarios para fácil acceso (ID de unidad como clave)
        diesel_map = {item['unidad_id']: item for item in consumo_diesel}
        urea_map = {item['unidad_id']: item for item in consumo_urea}
        aceite_map = {item['unidad_id']: item for item in consumo_aceite}

        # 3. Construir la estructura de datos final para la plantilla
        consumo_por_unidad = []
        for unidad in unidades:
            diesel_data = diesel_map.get(unidad.id, {})
            urea_data = urea_map.get(unidad.id, {})
            aceite_data = aceite_map.get(unidad.id, {})
            
            consumo_por_unidad.append({
                'unidad': unidad,
                'motor': diesel_data.get('total_motor', 0) or 0,
                'thermo': diesel_data.get('total_thermo', 0) or 0,
                'urea': urea_data.get('total_urea', 0) or 0,
                'aceite': aceite_data.get('total_aceite', 0) or 0,
            })
            
        # === NUEVA LÍNEA PARA ORDENAR LOS DATOS PARA EL GRÁFICO ===
        consumo_por_unidad.sort(key=lambda item: item['motor'] + item['thermo'], reverse=True)
            
        context['consumo_por_unidad'] = consumo_por_unidad
        return context


# -----------------------------------------------------------------------------
# 3. VISTAS PARA ADMINISTRADORES (CRUD COMPLETOS)
# -----------------------------------------------------------------------------

# ... (El resto de las vistas de Administrador no cambian) ...
# --- Gestión de Unidades ---
class UnidadListView(AdminRequiredMixin, ListView):
    model = Unidad
    template_name = 'generic_list.html'
    paginate_by = 25

    def get_queryset(self):
        queryset = super().get_queryset().order_by('nombre')
        
        # Leemos los parámetros de la URL
        search_query = self.request.GET.get('q')
        tipo_unidad = self.request.GET.get('tipo_unidad') # <-- LÍNEA NUEVA

        # Aplicamos filtro por nombre (como antes)
        if search_query:
            queryset = queryset.filter(nombre__icontains=search_query)
            
        # APLICAMOS EL NUEVO FILTRO POR TIPO DE UNIDAD
        if tipo_unidad:
            queryset = queryset.filter(tipo=tipo_unidad)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Unidades',
            'url_crear': 'unidad-create',
            'headers': ['#', 'Nombre', 'Marca/Modelo', 'Placas', 'Kilometraje', 'Tipo'],
            'url_detail_name': 'unidad-detail',
            'url_update_name': 'unidad-update',
            'url_delete_name': 'unidad-delete',
        })
        
        # Devolvemos los valores de los filtros a la plantilla
        context['search_query'] = self.request.GET.get('q', '')
        context['selected_tipo'] = self.request.GET.get('tipo_unidad', '') # <-- LÍNEA NUEVA
        
        return context
class UnidadDetailView(AdminRequiredMixin, DetailView):
    model = Unidad
    template_name = 'unidad_detail.html'

class UnidadCreateView(AdminRequiredMixin, CreateView):
    model = Unidad
    form_class = UnidadForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('unidad-list')
    extra_context = {'titulo': 'Registrar Nueva Unidad'}

class UnidadUpdateView(AdminRequiredMixin, UpdateView):
    model = Unidad
    form_class = UnidadForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('unidad-list')
    extra_context = {'titulo': 'Editar Unidad'}

# --- Gestión de Operadores ---
class OperadorListView(AdminRequiredMixin, ListView):
    model = Operador
    template_name = 'generic_list.html'
    extra_context = {'titulo': 'Operadores', 'url_crear': 'operador-create'}

class OperadorCreateView(AdminRequiredMixin, CreateView):
    model = Operador
    form_class = OperadorForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('operador-list')
    extra_context = {'titulo': 'Registrar Nuevo Operador'}

class OperadorUpdateView(AdminRequiredMixin, UpdateView):
    model = Operador
    form_class = OperadorForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('operador-list')
    extra_context = {'titulo': 'Editar Operador'}

# --- Gestión de Cargas de Diésel ---
class CargaDieselListView(AdminRequiredMixin, ListView):
    model = CargaDiesel
    template_name = 'generic_list.html'
    # COMENTA O ELIMINA LA SIGUIENTE LÍNEA:
    # extra_context = {'titulo': 'Cargas de Diésel', 'url_crear': 'cargadiesel-create'}
    
    # AÑADE ESTE MÉTODO:
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Cargas de Diésel',
            'url_crear': 'cargadiesel-create',
            'headers': ['Unidad', 'Fecha', 'Litros', 'Costo'],
            'url_update_name': 'cargadiesel-update',
            'url_delete_name': 'cargadiesel-delete',
        })
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            # Filtra por el nombre de la unidad relacionada
            queryset = queryset.filter(unidad__nombre__icontains=query)
        return queryset

class CargaDieselCreateView(AdminRequiredMixin, CreateView):
    model = CargaDiesel
    form_class = CargaDieselForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargadiesel-list')
    extra_context = {'titulo': 'Registrar Carga de Diésel'}

    # --- AÑADE ESTE MÉTODO ---
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Pasamos el usuario al form
        return kwargs

# --- Gestión de Cargas de Aceite ---
class CargaAceiteListView(AdminRequiredMixin, ListView):
    model = CargaAceite
    template_name = 'generic_list.html'

    def get_queryset(self):
        queryset = super().get_queryset().order_by('-fecha')
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        
        # --- LÍNEA AÑADIDA ---
        unidad_id = self.request.GET.get('unidad')

        if start_date_str and end_date_str:
            try:
                start_date = date.fromisoformat(start_date_str)
                end_date = date.fromisoformat(end_date_str)
                queryset = queryset.filter(fecha__date__range=[start_date, end_date])
            except (ValueError, TypeError):
                pass
        
        # --- BLOQUE AÑADIDO ---
        # Si se seleccionó una unidad en el filtro, se aplica aquí.
        if unidad_id:
            queryset = queryset.filter(unidad_id=unidad_id)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Cargas de Aceite',
            'url_crear': 'cargaaceite-create',
            'headers': ['#', 'Unidad', 'Fecha', 'Cantidad', 'Motivo'],
            'url_update_name': 'cargaaceite-update',
            'url_delete_name': 'cargaaceite-delete',
        })
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')

        # --- BLOQUE AÑADIDO ---
        # Esto es para que el campo "Unidad" recuerde la selección después de buscar.
        selected_unidad_id = self.request.GET.get('unidad')
        if selected_unidad_id:
            context['selected_unidad'] = Unidad.objects.filter(pk=selected_unidad_id).first()

        return context


class CargaAceiteCreateView(AdminRequiredMixin, CreateView):
    model = CargaAceite
    form_class = CargaAceiteForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargaaceite-list')
    extra_context = {'titulo': 'Registrar Carga de Aceite'}

# --- Gestión de Cargas de Urea ---
class CargaUreaListView(AdminRequiredMixin, ListView):
    model = CargaUrea
    template_name = 'generic_list.html'
    # COMENTA O ELIMINA LA SIGUIENTE LÍNEA:
    # extra_context = {'titulo': 'Cargas de Urea', 'url_crear': 'cargaurea-create'}

    # AÑADE ESTE MÉTODO:
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Cargas de Urea',
            'url_crear': 'cargaurea-create',
            'headers': ['Unidad', 'Fecha', 'Litros Cargados', 'Comentarios'],
            'url_update_name': 'cargaurea-update',
            'url_delete_name': 'cargaurea-delete',
        })
        return context
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            # Filtra por el nombre de la unidad relacionada
            queryset = queryset.filter(unidad__nombre__icontains=query)
        return queryset

class CargaUreaCreateView(AdminRequiredMixin, CreateView):
    model = CargaUrea
    form_class = CargaUreaForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargaurea-list')
    extra_context = {'titulo': 'Registrar Carga de Urea'}

# --- Gestión de Compras de Suministros ---
class CompraSuministroListView(AdminRequiredMixin, ListView):
    model = CompraSuministro
    template_name = 'generic_list.html'
    extra_context = {'titulo': 'Compras de Suministros', 'url_crear': 'comprasuministro-create'}

    # Añade este método para calcular el total
    def get_context_data(self, **kwargs):
        # Llama a la implementación base primero para obtener el contexto
        context = super().get_context_data(**kwargs)
        
        # Calcula la suma de todos los precios y añádela al contexto
        total_inversion = CompraSuministro.objects.aggregate(total=Sum('precio'))['total'] or 0
        context['total_inversion'] = total_inversion
        
        return context

class CompraSuministroCreateView(AdminRequiredMixin, CreateView):
    model = CompraSuministro
    form_class = CompraSuministroForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('comprasuministro-list')
    extra_context = {'titulo': 'Registrar Compra de Suministro'}


# -----------------------------------------------------------------------------
# 4. VISTAS PARA EL FLUJO DEL TÉCNICO
# -----------------------------------------------------------------------------
# --- VISTA ELIMINADA ---
# Se eliminó TecnicoDashboardView porque ya no es necesaria.


# --- Imports adicionales necesarios para esta sección ---
from django.views.generic.edit import FormView
from django.db import transaction
from django.contrib import messages
from django.forms import formset_factory # <-- AÑADIR ESTA IMPORTACIÓN


class TecnicoSeleccionarUnidadView(TecnicoRequiredMixin, ListView):
    """Paso 1: El técnico elige una unidad para iniciar el proceso."""
    model = Unidad
    template_name = 'tecnico_seleccionar_unidad.html'
    context_object_name = 'unidades'

class TecnicoChecklistView(TecnicoRequiredMixin, FormView):
    """Paso 2: Llenar el Checklist. Los datos se guardan en la sesión."""
    form_class = ChecklistInspeccionForm
    template_name = 'checklist_form.html'

    # --- MÉTODO MODIFICADO ---
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        context['titulo'] = f"Paso 1: Checklist para la Unidad {unidad}"
        
        # --- LÍNEA AÑADIDA ---
        # Pasamos el objeto 'unidad' completo a la plantilla
        context['unidad'] = unidad
        
        form = context['form']
        
        # (La lógica para estructurar el formulario no cambia)
        field_groups = {
            'Estructura Exterior': ['cristales', 'espejos', 'logos', 'num_economico', 'puertas', 'cofre', 'parrilla', 'defensas', 'faros', 'plafoneria', 'stops', 'direccionales', 'tapiceria', 'instrumentos', 'carroceria', 'piso', 'costados', 'escape', 'pintura', 'franjas', 'loderas', 'extintor', 'senalamientos', 'estado_general'],
            'Mecánica y Motor': ['motor', 'caja', 'diferenciales', 'suspension_delantera', 'suspension_trasera', 'fugas_combustible', 'fugas_aceite', 'estado_llantas', 'presion_llantas', 'purga_tanques', 'estado_balatas', 'amortiguadores_delanteros', 'amortiguadores_traseros', 'rines_aluminio', 'mangueras_servicio', 'tarjeta_llave', 'revision_fusibles', 'revision_luces', 'revision_cable_7vias', 'revision_fuga_aire']
        }
        structured_form = {}
        for group_name, field_list in field_groups.items():
            structured_form[group_name] = []
            for field_name in field_list:
                if field_name in form.fields:
                    obs_field_name = f"{field_name}_obs"
                    structured_form[group_name].append({'status': form[field_name], 'observation': form[obs_field_name] if obs_field_name in form.fields else None})
        context['structured_form'] = structured_form
        return context

    def get_initial(self):
        # Limpiamos datos de un proceso anterior para evitar conflictos
        self.request.session.pop('checklist_data', None)
        self.request.session.pop('llantas_data', None) # <-- Limpiar sesión de llantas
        self.request.session.pop('diesel_data', None)
        return {'unidad': get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])}

    def form_valid(self, form):
        # Guardamos los datos del formulario en la sesión en lugar de la BD
        checklist_data = form.cleaned_data
        # Convertimos objetos (como Operador) a su ID para que se puedan guardar en la sesión
        checklist_data['operador_id'] = checklist_data.pop('operador').id
        del checklist_data['unidad'] # La unidad ya la tenemos en la URL
        
        self.request.session['checklist_data'] = checklist_data
        # --- LÍNEA MODIFICADA ---
        # Ahora redirige al nuevo paso de llantas
        return redirect('tecnico-proceso-llantas', unidad_pk=self.kwargs['unidad_pk'])

# --- INICIO DE VISTA AÑADIDA ---
class TecnicoProcesoLlantasView(TecnicoRequiredMixin, TemplateView):
    """Paso 3: Llenar el formato de Llantas. Lógica de POST corregida."""
    template_name = 'llantas_form.html'

    def dispatch(self, request, *args, **kwargs):
        if 'checklist_data' not in request.session:
            messages.error(request, 'Error: Debe completar el checklist antes de continuar.')
            return redirect('tecnico-seleccionar-unidad')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        # Este método ahora solo se encarga de mostrar el formulario la primera vez (GET)
        context = super().get_context_data(**kwargs)
        unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        
        posiciones_llantas = [
            {'posicion': 'Posición 1'}, {'posicion': 'Posición 2'},
            {'posicion': 'Posición 3'}, {'posicion': 'Posición 4'},
            {'posicion': 'Posición 5'}, {'posicion': 'Posición 6'},
        ]
        
        LlantaFormSet = formset_factory(LlantaDetalleForm, extra=6, max_num=6)
        
        context['titulo'] = f"Paso 2: Formato de Llantas para {unidad}"
        context['unidad'] = unidad
        # Se inicializan los formularios vacíos, pasando la unidad para la validación de KM
        context['km_form'] = LlantasKmForm(initial={'km': unidad.km_actual}, unidad=unidad)
        context['formset'] = LlantaFormSet(initial=posiciones_llantas, prefix='llantas')
        context['fecha_actual'] = date.today()
        return context

    def post(self, request, *args, **kwargs):
        # Este método ahora maneja la lógica de envío (POST)
        unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        LlantaFormSet = formset_factory(LlantaDetalleForm, extra=6, max_num=6)

        # 1. Creamos las instancias de los formularios CON los datos enviados (request.POST)
        km_form = LlantasKmForm(request.POST, unidad=unidad)
        formset = LlantaFormSet(request.POST, prefix='llantas')

        # 2. Validamos ambos formularios
        if km_form.is_valid() and formset.is_valid():
            # Si TODO es válido, guardamos en sesión y redirigimos
            llantas_data = {
                'km_data': km_form.cleaned_data,
                'formset_data': formset.cleaned_data
            }
            request.session['llantas_data'] = llantas_data
            return redirect('tecnico-proceso-diesel', unidad_pk=self.kwargs['unidad_pk'])
        else:
            # --- BLOQUE CORREGIDO ---
            # 3. Si hay errores, volvemos a renderizar la página
            # PERO pasamos los formularios inválidos que ya contienen los datos del usuario y los errores.
            # NO llamamos a get_context_data() aquí.
            messages.error(request, 'Por favor, corrija los errores marcados en el formulario.')
            
            # Construimos el contexto manualmente para la respuesta de error
            context = {
                'titulo': f"Paso 2: Formato de Llantas para {unidad}",
                'unidad': unidad,
                'km_form': km_form,     # Usamos el formulario con datos y errores
                'formset': formset,   # Usamos el formset con datos y errores
                'fecha_actual': date.today()
            }
            return self.render_to_response(context)

class TecnicoProcesoDieselView(TecnicoRequiredMixin, FormView):
    """Paso 4: Llenar formulario de Diésel. Los datos se guardan en la sesión."""
    form_class = CargaDieselForm
    template_name = 'generic_form.html'

    # --- MÉTODO CORREGIDO Y UNIFICADO ---
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['unidad'] = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        return kwargs

    def dispatch(self, request, *args, **kwargs):
        # --- LÓGICA MODIFICADA ---
        # Ahora valida que los dos pasos anteriores se hayan completado
        if 'checklist_data' not in request.session or 'llantas_data' not in request.session:
            messages.error(request, 'Error: Debe completar el checklist y el formato de llantas antes de continuar.')
            return redirect('tecnico-seleccionar-unidad')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        # --- TEXTO MODIFICADO ---
        context['titulo'] = f"Paso 3: Carga de Diésel para {unidad}"
        return context

    def get_initial(self):
        unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        
        ultima_carga = CargaDiesel.objects.filter(unidad=unidad).order_by('-fecha').first()
        
        cinchos_previos = ""
        if ultima_carga and ultima_carga.cinchos_actuales:
            cinchos_previos = ultima_carga.cinchos_actuales

        # Usamos el KM del formulario de llantas que ya está en la sesión
        km_llantas = self.request.session.get('llantas_data', {}).get('km_data', {}).get('km', unidad.km_actual)
        
        return {
            'unidad': unidad,
            'operador': get_object_or_404(Operador, pk=self.request.session['checklist_data']['operador_id']),
            'km_actual': km_llantas, # <-- Usamos el KM del paso anterior
            'persona_relleno': self.request.user.get_full_name() or self.request.user.username,
            'cinchos_anteriores': cinchos_previos,
        }
        
    def form_valid(self, form):
        diesel_data = form.cleaned_data
        diesel_data['operador_id'] = diesel_data.pop('operador').id
        if 'unidad' in diesel_data:
            del diesel_data['unidad']
        
        for key, value in diesel_data.items():
            if isinstance(value, Decimal):
                diesel_data[key] = str(value)

        self.request.session['diesel_data'] = diesel_data
        return redirect('tecnico-proceso-urea', unidad_pk=self.kwargs['unidad_pk'])


class TecnicoProcesoUreaView(TecnicoRequiredMixin, FormView):
    """Paso 5: Llenar formulario de Urea y guardar todo el proceso."""
    form_class = CargaUreaForm
    template_name = 'generic_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unidad = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        context['titulo'] = f"Paso 4: Carga de Urea y Finalizar Proceso para {unidad}"
        return context

    def get_initial(self):
        unidad_obj = get_object_or_404(Unidad, pk=self.kwargs['unidad_pk'])
        return {'unidad': unidad_obj}

    def form_valid(self, form):
        unidad_pk = self.kwargs['unidad_pk']
        unidad = get_object_or_404(Unidad, pk=unidad_pk)
        
        checklist_data = self.request.session.get('checklist_data')
        llantas_data = self.request.session.get('llantas_data')
        diesel_data = self.request.session.get('diesel_data')
        urea_data = form.cleaned_data

        if not all([checklist_data, llantas_data, diesel_data]):
            messages.error(self.request, "La sesión ha expirado o faltan datos. Inicie el proceso de nuevo.")
            return redirect('tecnico-seleccionar-unidad')
        
        # --- Corrección de Tipos de Dato (Diésel) ---
        campos_decimales_diesel = ['lts_diesel', 'lts_thermo', 'hrs_thermo']
        for campo in campos_decimales_diesel:
            if campo in diesel_data and diesel_data[campo]:
                try:
                    diesel_data[campo] = Decimal(diesel_data[campo])
                except (ValueError, TypeError):
                    messages.error(self.request, f"Error de formato en el campo '{campo}'.")
                    return redirect('tecnico-seleccionar-unidad')

        # --- Corrección de Tipos de Dato (Llantas) ---
        if 'formset_data' in llantas_data:
            for detalle_llanta in llantas_data['formset_data']:
                if 'mm' in detalle_llanta and detalle_llanta['mm']:
                    try:
                        detalle_llanta['mm'] = Decimal(detalle_llanta['mm'])
                    except (ValueError, TypeError):
                        messages.error(self.request, f"Error de formato en el campo 'mm' de las llantas.")
                        return redirect('tecnico-seleccionar-unidad')

        try:
            with transaction.atomic():
                # 1. Guardado del Checklist
                operador_checklist = get_object_or_404(Operador, pk=checklist_data.pop('operador_id'))
                ChecklistInspeccion.objects.create(
                    unidad=unidad, operador=operador_checklist, tecnico=self.request.user, **checklist_data
                )
                
                # 2. Guardado de Inspección de Llantas
                km_llantas = llantas_data['km_data']['km']
                inspeccion_llantas = LlantasInspeccion.objects.create(
                    unidad=unidad, tecnico=self.request.user, km=km_llantas
                )
                
                # --- INICIO DE LA LÓGICA CORREGIDA ---
                # Iteramos sobre los datos del formset y solo guardamos las filas
                # que el usuario realmente llenó (verificando que 'mm' tenga un valor).
                for detalle_data in llantas_data['formset_data']:
                    if detalle_data and detalle_data.get('mm'):
                        LlantaDetalle.objects.create(inspeccion=inspeccion_llantas, **detalle_data)
                # --- FIN DE LA LÓGICA CORREGIDA ---

                # 3. Lógica de cálculo de rendimiento
                ultima_carga_diesel = CargaDiesel.objects.filter(unidad=unidad).order_by('-fecha').first()
                km_recorridos = 0
                km_actual_diesel = int(diesel_data.get('km_actual', 0))
                
                if ultima_carga_diesel and km_actual_diesel > ultima_carga_diesel.km_actual:
                    km_recorridos = km_actual_diesel - ultima_carga_diesel.km_actual

                if km_recorridos > 0:
                    lts_diesel = diesel_data.get('lts_diesel', Decimal('0'))
                    if lts_diesel and lts_diesel > 0:
                        diesel_data['rendimiento'] = Decimal(km_recorridos) / lts_diesel

                    lts_urea = urea_data.get('litros_cargados', Decimal('0'))
                    if lts_urea and lts_urea > 0:
                        urea_data['rendimiento'] = Decimal(km_recorridos) / lts_urea
                
                # 4. Guardado de Diésel
                operador_diesel = get_object_or_404(Operador, pk=diesel_data.pop('operador_id'))
                CargaDiesel.objects.create(
                    unidad=unidad, operador=operador_diesel, **diesel_data
                )

                # 5. Guardado de Urea
                if 'unidad' in urea_data:
                    del urea_data['unidad']
                CargaUrea.objects.create(
                    unidad=unidad, **urea_data
                )

                # 6. Actualización del Kilometraje de la unidad
                unidad.km_actual = km_actual_diesel
                unidad.save()
            
            self.request.session.pop('checklist_data', None)
            self.request.session.pop('llantas_data', None)
            self.request.session.pop('diesel_data', None)
            messages.success(self.request, f"Proceso para la unidad {unidad} completado y guardado exitosamente.")

        except Exception as e:
            # Ahora puedes volver a poner el mensaje de error amigable
            messages.error(self.request, f"Ocurrió un error inesperado al guardar los datos: {e}. Por favor, intente de nuevo.")

        return redirect('tecnico-seleccionar-unidad')

class DashboardGraficasView(AdminRequiredMixin, ListView):
    model = Unidad
    template_name = 'dashboard_graficas.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # --- Lógica de Fechas (sin cambios) ---
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        if start_date_str and end_date_str:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
        else:
            today = date.today()
            last_sunday = today - timedelta(days=(today.weekday() + 1))
            last_monday = last_sunday - timedelta(days=6)
            start_date = last_monday
            end_date = last_sunday
            
        context['start_date'] = start_date
        context['end_date'] = end_date

        # --- Cálculos para las Tarjetas ---
        
        # 1. Diésel Comprado en el rango de fechas
        total_diesel_comprado_rango = CompraSuministro.objects.filter(
            tipo_suministro='DIESEL',
            fecha_compra__range=[start_date, end_date]
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        context['total_diesel_comprado_rango'] = total_diesel_comprado_rango

        # 2. Consumo de Diésel en el rango de fechas
        consumo_diesel = CargaDiesel.objects.filter(fecha__date__range=[start_date, end_date]) \
            .values('unidad_id') \
            .annotate(total_motor=Sum('lts_diesel'), total_thermo=Sum('lts_thermo'))
        
        diesel_map = {item['unidad_id']: item for item in consumo_diesel}

        # --- Procesamiento de datos para las gráficas ---
        unidades = Unidad.objects.all()
        consumo_por_unidad = []
        for unidad in unidades:
            diesel_data = diesel_map.get(unidad.id, {})
            motor = diesel_data.get('total_motor', 0) or 0
            thermo = diesel_data.get('total_thermo', 0) or 0
            
            if motor > 0 or thermo > 0:
                # Usamos el campo 'nombre' de la unidad
                consumo_por_unidad.append({
                    'unidad': unidad.nombre, 
                    'motor': motor, 
                    'thermo': thermo
                })
        
        # Datos para la gráfica de Motores (ordenados de mayor a menor)
        motor_chart_data = sorted(
            [item for item in consumo_por_unidad if item['motor'] > 0],
            key=lambda x: x['motor'], reverse=True
        )
        context['motor_chart_data'] = motor_chart_data

        # Datos para la gráfica de Thermos (ordenados de mayor a menor)
        thermo_chart_data = sorted(
            [item for item in consumo_por_unidad if item['thermo'] > 0],
            key=lambda x: x['thermo'], reverse=True
        )
        context['thermo_chart_data'] = thermo_chart_data

        # 3. Totales para las tarjetas de consumo
        total_motor_rango = sum(item['motor'] for item in consumo_por_unidad)
        total_thermo_rango = sum(item['thermo'] for item in consumo_por_unidad)
        context['total_motor_rango'] = total_motor_rango
        context['total_thermo_rango'] = total_thermo_rango
        context['total_general_rango'] = total_motor_rango + total_thermo_rango

        return context
    
class UnidadDetailRendimientoView(AdminRequiredMixin, DetailView):
    model = Unidad
    template_name = 'unidad_rendimiento_detail.html'
    context_object_name = 'unidad'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unidad = self.get_object()
        context['cargas_diesel'] = CargaDiesel.objects.filter(unidad=unidad).order_by('-fecha')
        context['cargas_urea'] = CargaUrea.objects.filter(unidad=unidad).order_by('-fecha')
        return context
    
    
# --- Gestión de Checklists ---
class ChecklistListView(AdminRequiredMixin, ListView):
    model = ChecklistInspeccion
    template_name = 'generic_list.html'
    context_object_name = 'object_list'
    ordering = ['-fecha']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Registros de Checklist',
            'url_crear': 'checklist-create',
            'headers': ['Unidad', 'Operador', 'Técnico', 'Fecha'],
            'fields': ['unidad', 'operador', 'tecnico', 'fecha'],
            'url_detail_name': 'checklist-detail',
            'url_update_name': 'checklist-update',
            'url_delete_name': 'checklist-delete',
        })
        return context
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            # Filtra por el nombre de la unidad relacionada
            queryset = queryset.filter(unidad__nombre__icontains=query)
        return queryset

class ChecklistDetailView(AdminRequiredMixin, DetailView):
    model = ChecklistInspeccion
    template_name = 'checklist_detail.html'

class ChecklistCreateView(AdminRequiredMixin, CreateView):
    model = ChecklistInspeccion
    form_class = ChecklistInspeccionForm
    template_name = 'checklist_form.html' # Reutilizamos el form del técnico
    success_url = reverse_lazy('checklist-list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = "Crear Nuevo Registro de Checklist"
        # Lógica para estructurar el formulario (copiada de la vista del técnico)
        form = context['form']
        field_groups = {
            'Estructura Exterior': ['cristales', 'espejos', 'logos', 'num_economico', 'puertas', 'cofre', 'parrilla', 'defensas', 'faros', 'plafoneria', 'stops', 'direccionales', 'tapiceria', 'instrumentos', 'carroceria', 'piso', 'costados', 'escape', 'pintura', 'franjas', 'loderas', 'extintor', 'senalamientos', 'estado_general'],
            'Mecánica y Motor': ['motor', 'caja', 'diferenciales', 'suspension_delantera', 'suspension_trasera', 'fugas_combustible', 'fugas_aceite', 'estado_llantas', 'presion_llantas', 'purga_tanques', 'estado_balatas', 'amortiguadores_delanteros', 'amortiguadores_traseros', 'rines_aluminio', 'mangueras_servicio', 'tarjeta_llave', 'revision_fusibles', 'revision_luces', 'revision_cable_7vias', 'revision_fuga_aire']
        }
        structured_form = {}
        for group_name, field_list in field_groups.items():
            structured_form[group_name] = []
            for field_name in field_list:
                if field_name in form.fields:
                    obs_field_name = f"{field_name}_obs"
                    structured_form[group_name].append({'status': form[field_name], 'observation': form[obs_field_name] if obs_field_name in form.fields else None})
        context['structured_form'] = structured_form
        return context

    def form_valid(self, form):
        form.instance.tecnico = self.request.user
        messages.success(self.request, "Checklist creado exitosamente.")
        return super().form_valid(form)

class ChecklistUpdateView(AdminRequiredMixin, UpdateView):
    model = ChecklistInspeccion
    form_class = ChecklistInspeccionForm
    template_name = 'checklist_form.html'
    success_url = reverse_lazy('checklist-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f"Editar Checklist de {self.object.unidad}"
        # (Lógica para estructurar el formulario, igual que en CreateView)
        form = context['form']
        field_groups = {
            'Estructura Exterior': ['cristales', 'espejos', 'logos', 'num_economico', 'puertas', 'cofre', 'parrilla', 'defensas', 'faros', 'plafoneria', 'stops', 'direccionales', 'tapiceria', 'instrumentos', 'carroceria', 'piso', 'costados', 'escape', 'pintura', 'franjas', 'loderas', 'extintor', 'senalamientos', 'estado_general'],
            'Mecánica y Motor': ['motor', 'caja', 'diferenciales', 'suspension_delantera', 'suspension_trasera', 'fugas_combustible', 'fugas_aceite', 'estado_llantas', 'presion_llantas', 'purga_tanques', 'estado_balatas', 'amortiguadores_delanteros', 'amortiguadores_traseros', 'rines_aluminio', 'mangueras_servicio', 'tarjeta_llave', 'revision_fusibles', 'revision_luces', 'revision_cable_7vias', 'revision_fuga_aire']
        }
        structured_form = {}
        for group_name, field_list in field_groups.items():
            structured_form[group_name] = []
            for field_name in field_list:
                if field_name in form.fields:
                    obs_field_name = f"{field_name}_obs"
                    structured_form[group_name].append({'status': form[field_name], 'observation': form[obs_field_name] if obs_field_name in form.fields else None})
        context['structured_form'] = structured_form
        return context
        
    def form_valid(self, form):
        messages.success(self.request, "Checklist actualizado exitosamente.")
        return super().form_valid(form)

class ChecklistDeleteView(AdminRequiredMixin, DeleteView):
    model = ChecklistInspeccion
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('checklist-list')
    
    def form_valid(self, form):
        messages.success(self.request, "Registro de checklist eliminado exitosamente.")
        return super().form_valid(form)


class LlantasInspeccionListView(AdminRequiredMixin, ListView):
    model = LlantasInspeccion
    template_name = 'generic_list.html'
    context_object_name = 'object_list'
    ordering = ['-fecha']

    # --- MÉTODO CORREGIDO (SIN DUPLICADOS) ---
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'titulo': 'Inspecciones de Llantas',
            'url_crear': 'llantas-create',
            'headers': ['Unidad', 'Kilometraje', 'Técnico', 'Fecha'],
            'fields': ['unidad', 'km', 'tecnico', 'fecha'],
            'url_detail_name': 'llantas-detail',
            'url_update_name': 'llantas-update',
            'url_delete_name': 'llantas-delete',
        })
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            # Filtra por el nombre de la unidad relacionada
            queryset = queryset.filter(unidad__nombre__icontains=query)
        return queryset

class LlantasInspeccionDetailView(AdminRequiredMixin, DetailView):
    model = LlantasInspeccion
    template_name = 'llantas_inspeccion_detail.html'

class LlantasInspeccionCreateView(AdminRequiredMixin, CreateView):
    model = LlantasInspeccion
    form_class = LlantasInspeccionForm
    template_name = 'llantas_inspeccion_form.html'
    success_url = reverse_lazy('llantas-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = "Crear Inspección de Llantas"
        
        LlantaDetalleFormSet = inlineformset_factory(
            LlantasInspeccion, LlantaDetalle, form=LlantaDetalleForm,
            extra=6, max_num=6, can_delete=False
        )

        if self.request.POST:
            context['formset'] = LlantaDetalleFormSet(self.request.POST, prefix='llantas')
        else:
            initial_data = [{'posicion': f'Posición {i}'} for i in range(1, 7)]
            context['formset'] = LlantaDetalleFormSet(prefix='llantas', initial=initial_data)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                form.instance.tecnico = self.request.user
                self.object = form.save()
                formset.instance = self.object
                formset.save()
            
            messages.success(self.request, "Inspección de llantas creada exitosamente.")
            return redirect(self.success_url)
        else:
            return self.form_invalid(form)
        
class LlantasInspeccionUpdateView(AdminRequiredMixin, UpdateView):
    model = LlantasInspeccion
    form_class = LlantasInspeccionForm
    template_name = 'llantas_inspeccion_form.html'
    success_url = reverse_lazy('llantas-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f"Editar Inspección de Llantas de {self.object.unidad}"
        
        # Definimos el formset con extra=0 para que solo cargue los existentes
        LlantaDetalleFormSet = inlineformset_factory(
            LlantasInspeccion, LlantaDetalle, form=LlantaDetalleForm,
            extra=0, max_num=6, can_delete=True
        )

        if self.request.POST:
            context['formset'] = LlantaDetalleFormSet(self.request.POST, instance=self.object, prefix='llantas')
        else:
            # === LA CLAVE ESTÁ AQUÍ ===
            # Cuando no hay POST, le pasamos la instancia para que cargue los datos existentes.
            # Pero además, si no hay datos guardados, queremos 6 formularios vacíos iniciales
            # para que el usuario pueda añadir llantas si es una inspección "vacía".
            
            # Contar cuántas llantas_detalles tiene esta inspección
            num_detalles_existentes = self.object.llantadetalle_set.count()
            
            # Si no hay detalles de llantas guardados, entonces creamos 6 formularios "iniciales"
            # para que el usuario pueda empezar a añadir.
            if num_detalles_existentes == 0:
                # Modificamos el formset para que temporalmente tenga 6 extras para este caso
                LlantaDetalleFormSet = inlineformset_factory(
                    LlantasInspeccion, LlantaDetalle, form=LlantaDetalleForm,
                    extra=6, max_num=6, can_delete=False # Can_delete en False para nuevas entradas
                )
                initial_data = [{'posicion': f'Posición {i}'} for i in range(1, 7)]
                context['formset'] = LlantaDetalleFormSet(instance=self.object, prefix='llantas', initial=initial_data)
            else:
                # Si ya hay detalles, simplemente cargamos los existentes (extra=0 se encarga)
                context['formset'] = LlantaDetalleFormSet(instance=self.object, prefix='llantas')
        return context

    # El método form_valid no cambia
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                formset.instance = self.object
                formset.save()
            messages.success(self.request, "Inspección de llantas actualizada exitosamente.")
            return redirect(self.success_url)
        else:
            return self.form_invalid(form)


class LlantasInspeccionDeleteView(AdminRequiredMixin, DeleteView):
    model = LlantasInspeccion
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('llantas-list')

    def form_valid(self, form):
        messages.success(self.request, "Inspección de llantas eliminada exitosamente.")
        return super().form_valid(form)
    
class CargaDieselUpdateView(AdminRequiredMixin, UpdateView):
    model = CargaDiesel
    form_class = CargaDieselForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargadiesel-list')
    extra_context = {'titulo': 'Editar Carga de Diésel'}

class CargaDieselDeleteView(AdminRequiredMixin, DeleteView):
    model = CargaDiesel
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('cargadiesel-list')
    extra_context = {'titulo': 'Eliminar Carga de Diésel'}

# --- Vistas para Carga de Aceite ---
class CargaAceiteUpdateView(AdminRequiredMixin, UpdateView):
    model = CargaAceite
    form_class = CargaAceiteForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargaaceite-list')
    extra_context = {'titulo': 'Editar Carga de Aceite'}

class CargaAceiteDeleteView(AdminRequiredMixin, DeleteView):
    model = CargaAceite
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('cargaaceite-list')
    extra_context = {'titulo': 'Eliminar Carga de Aceite'}

# --- Vistas para Carga de Urea ---
class CargaUreaUpdateView(AdminRequiredMixin, UpdateView):
    model = CargaUrea
    form_class = CargaUreaForm
    template_name = 'generic_form.html'
    success_url = reverse_lazy('cargaurea-list')
    extra_context = {'titulo': 'Editar Carga de Urea'}

class CargaUreaDeleteView(AdminRequiredMixin, DeleteView):
    model = CargaUrea
    template_name = 'generic_confirm_delete.html'
    success_url = reverse_lazy('cargaurea-list')
    extra_context = {'titulo': 'Eliminar Carga de Urea'}
    

# ... (después de tus vistas de Admin y las vistas del flujo de inicio) ...

# -----------------------------------------------------------------------------
# 5. VISTAS PARA EL FLUJO DEL ENCARGADO (ETAPA 2) - NUEVAS
# -----------------------------------------------------------------------------

class EncargadoPendientesListView(EncargadoRequiredMixin, ListView):
    """Página principal del Encargado: Muestra unidades pendientes de carga."""
    model = ProcesoCarga
    template_name = 'encargado_pendientes_list.html'
    context_object_name = 'procesos_pendientes'

    def get_queryset(self):
        # Filtra solo los procesos que no han sido completados
        return ProcesoCarga.objects.filter(status='PENDIENTE').select_related('unidad', 'tecnico_inicia')

class EncargadoProcesoDieselView(EncargadoRequiredMixin, FormView):
    """Paso 4: El Encargado llena los datos del Diésel."""
    form_class = CargaDieselForm
    template_name = 'generic_form.html'

    def get_proceso(self):
        # Helper para obtener el proceso actual
        return get_object_or_404(ProcesoCarga, pk=self.kwargs['proceso_pk'], status='PENDIENTE')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        proceso = self.get_proceso()
        context['titulo'] = f"Paso 3: Carga de Diésel para {proceso.unidad.nombre}"
        context['url_cancelar'] = reverse('encargado-pendientes-list')
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        proceso = self.get_proceso()
        kwargs['unidad'] = proceso.unidad # Pasa la unidad al form para validaciones
        return kwargs

    def get_initial(self):
        proceso = self.get_proceso()
        unidad = proceso.unidad
        ultima_carga = CargaDiesel.objects.filter(unidad=unidad).order_by('-fecha').first()
        cinchos_previos = ultima_carga.cinchos_actuales if ultima_carga else ""

        return {
            'unidad': unidad,
            'operador': proceso.checklist.operador,
            'km_actual': proceso.inspeccion_llantas.km,
            'persona_relleno': self.request.user.get_full_name() or self.request.user.username,
            'cinchos_anteriores': cinchos_previos,
        }

    def form_valid(self, form):
        diesel_data = form.cleaned_data
        diesel_data['operador_id'] = diesel_data.pop('operador').id
        if 'unidad' in diesel_data: del diesel_data['unidad']
        
        # Serializar datos para la sesión
        for key, value in diesel_data.items():
            if isinstance(value, Decimal):
                diesel_data[key] = str(value)

        # Guarda en sesión usando una clave única por proceso
        self.request.session[f'diesel_data_proceso_{self.kwargs["proceso_pk"]}'] = diesel_data
        return redirect('encargado-proceso-urea', proceso_pk=self.kwargs['proceso_pk'])

class EncargadoProcesoUreaView(EncargadoRequiredMixin, FormView):
    """Paso 5 y Final: El Encargado llena Urea y completa el proceso."""
    form_class = CargaUreaForm
    template_name = 'generic_form.html'

    def get_proceso(self):
        return get_object_or_404(ProcesoCarga, pk=self.kwargs['proceso_pk'], status='PENDIENTE')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        proceso = self.get_proceso()
        context['titulo'] = f"Paso 4: Carga de Urea y Finalizar para {proceso.unidad.nombre}"
        context['url_cancelar'] = reverse('encargado-pendientes-list')
        return context
    
    def get_initial(self):
        return {'unidad': self.get_proceso().unidad}

    def form_valid(self, form):
        proceso = self.get_proceso()
        unidad = proceso.unidad
        
        diesel_data = self.request.session.get(f'diesel_data_proceso_{proceso.pk}')
        urea_data = form.cleaned_data

        if not diesel_data:
            messages.error(self.request, "Sesión expirada. Faltan datos de diésel.")
            return redirect('encargado-proceso-diesel', proceso_pk=proceso.pk)

        # Convertir datos de sesión a sus tipos correctos
        for campo in ['lts_diesel', 'lts_thermo', 'hrs_thermo', 'costo']:
            if campo in diesel_data and diesel_data[campo]:
                diesel_data[campo] = Decimal(diesel_data[campo])

        try:
            with transaction.atomic():
                # --- Guardado de Diésel ---
                operador = get_object_or_404(Operador, pk=diesel_data.pop('operador_id'))
                # Lógica de rendimiento
                km_actual_diesel = int(diesel_data.get('km_actual', 0))
                # ... (resto de tu lógica de rendimiento aquí) ...
                
                carga_diesel_obj = CargaDiesel.objects.create(unidad=unidad, operador=operador, **diesel_data)
                
                # --- Guardado de Urea ---
                carga_urea_obj = None
                if urea_data.get('litros_cargados', 0) > 0:
                    del urea_data['unidad']
                    carga_urea_obj = CargaUrea.objects.create(unidad=unidad, **urea_data)

                # --- Actualización del ProcesoCarga ---
                proceso.carga_diesel = carga_diesel_obj
                proceso.carga_urea = carga_urea_obj
                proceso.encargado_finaliza = self.request.user
                proceso.fecha_fin = datetime.now()
                proceso.status = 'COMPLETADO'
                proceso.save()
            
            self.request.session.pop(f'diesel_data_proceso_{proceso.pk}', None)
            messages.success(self.request, f"Proceso para {unidad.nombre} finalizado.")
        except Exception as e:
            messages.error(self.request, f"Error inesperado al finalizar: {e}.")

        return redirect('encargado-pendientes-list')
    
@login_required
def download_checklist_excel(request, pk):
    """
    Genera y sirve un archivo Excel con el detalle de un Checklist de Inspección.
    """
    if not es_admin(request.user):
        raise PermissionDenied("No tiene permiso para exportar este reporte.")

    # 1. Obtener el objeto Checklist
    checklist = get_object_or_404(ChecklistInspeccion.objects.select_related(
        'unidad', 'operador', 'tecnico'
    ), pk=pk)

    # 2. Crear la respuesta HTTP con el tipo de contenido de Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # Define el nombre del archivo que se descargará
    filename = f"checklist_{checklist.unidad.nombre.replace(' ', '_')}_{checklist.fecha.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # 3. Crear el libro de trabajo y la hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle del Checklist"

    # 4. Definir estilos para los encabezados
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center')

    # 5. Agregar la información general del checklist
    ws['A1'] = 'Reporte de Checklist de Inspección'
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    ws.append([]) # Fila en blanco
    ws.append(['Unidad:', checklist.unidad.nombre])
    ws.append(['Operador:', str(checklist.operador)])
    ws.append(['Técnico:', checklist.tecnico.get_full_name() or checklist.tecnico.username])
    ws.append(['Fecha de Inspección:', checklist.fecha.strftime('%Y-%m-%d %H:%M:%S')])
    ws.append([]) # Fila en blanco

    # 6. Agregar la cabecera de la tabla de detalles
    header = ['Componente', 'Estado', 'Observaciones']
    ws.append(header)
    for cell in ws[ws.max_row]: # Aplica estilo a la última fila (la cabecera)
        cell.font = bold_font

    # 7. Iterar sobre los campos del modelo para poblar los datos
    # (Usando una lógica similar a la de la vista de creación/edición)
    field_groups = {
        'Estructura Exterior': ['cristales', 'espejos', 'logos', 'num_economico', 'puertas', 'cofre', 'parrilla', 'defensas', 'faros', 'plafoneria', 'stops', 'direccionales', 'tapiceria', 'instrumentos', 'carroceria', 'piso', 'costados', 'escape', 'pintura', 'franjas', 'loderas', 'extintor', 'senalamientos', 'estado_general'],
        'Mecánica y Motor': ['motor', 'caja', 'diferenciales', 'suspension_delantera', 'suspension_trasera', 'fugas_combustible', 'fugas_aceite', 'estado_llantas', 'presion_llantas', 'purga_tanques', 'estado_balatas', 'amortiguadores_delanteros', 'amortiguadores_traseros', 'rines_aluminio', 'mangueras_servicio', 'tarjeta_llave', 'revision_fusibles', 'revision_luces', 'revision_cable_7vias', 'revision_fuga_aire']
    }

    for group_name, field_list in field_groups.items():
        # Añadir una fila para el título del grupo
        ws.append([group_name])
        ws.merge_cells(f'A{ws.max_row}:C{ws.max_row}')
        ws.cell(row=ws.max_row, column=1).font = bold_font
        
        for field_name in field_list:
            field_obj = ChecklistInspeccion._meta.get_field(field_name)
            # Extraer el valor, la observación y el nombre amigable
            status = getattr(checklist, field_name, '')
            observation = getattr(checklist, f"{field_name}_obs", '')
            verbose_name = field_obj.verbose_name.title()
            
            # Añadir la fila con los datos
            ws.append([verbose_name, status, observation])

    # 8. Ajustar el ancho de las columnas para mejor legibilidad
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50

    # 9. Guardar el libro de trabajo en la respuesta HTTP
    wb.save(response)

    return response